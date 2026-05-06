"""
EMG Stimulation Artefact Suppression — Interactive Benchmark GUI  v3
Ayush Ram · Thesis A&B · University of Sydney

Algorithms (9 total):
  1. Blanking              §2.3
  2. Fixed template        §2.6.1
  3. EWMA template         §2.6.2
  4. Dual-ch / DESTD       §2.5.1
  5. GSO                   §2.5.2
  6. LMS                   §2.7.1
  7. ε-NLMS                §2.7.1
  8. RLS                   §2.7.1
  9. CEEMDAN               §2.7.2

METRIC MODES (selectable in sidebar):
  ── Current (Inter-stim)  Your existing approach: metrics on inter-stimulus
                           segments only, 30 ms guard excluded. Fair cross-
                           method comparison.

  ── Wang/Chen 2021/2023   SNR = 10·log10(E[e²]/E[(y-e)²])
                           NRMSE = rmse / std(clean)
                           Full signal, clean-referenced.
                           Matches Wang (GS-APEF/LMS) & Chen (GSO) papers.

  ── Sennels 1997          MRI_y = 10·log10(E[v²]/E[y²])
                           PR    = 10·log10(E[x²]/E[y²])
                           Matches adaptive filter (LMS/RLS/ε-NLMS) paper.

  ── ASR (Liu/Limnuson)    ASR = 20·log10(pp_art / pp_residual)
                           Peak-to-peak artefact suppression ratio.
                           Matches CMOS/FPGA template papers.

  ── Mandrile 2003         ARV_norm = mean|art_seg| / mean|mwave_seg| × 100%
                           Matches blanking/spatial characterisation paper.

Run:   python emg_gui_v3.py
Deps:  numpy  scipy  matplotlib  tkinter
"""

import os, time, warnings
import numpy as np
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.gridspec import GridSpec
from scipy import signal as sp
from scipy.signal import hilbert
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

warnings.filterwarnings("ignore")

# ─── palette ──────────────────────────────────────────────────────────────────
BG      = "#0f1117"
SURFACE = "#222633"
CARD    = "#20232f"
BORDER  = "#2e3347"
TEXT    = "#e8eaf0"
MUTED   = "#6b7280"
ACCENT  = "#6c8fff"
ACCENT2 = "#34d399"

ALGO_COLORS = {
    "blank"   : "#6b7280",
    "template": "#94a3b8",
    "ewma"    : "#38bdf8",
    "destd"   : "#6c8fff",
    "gso"     : "#34d399",
    "lms"     : "#a78bfa",
    "enlms"   : "#c084fc",
    "rls"     : "#f97316",
    "ceemdan" : "#fb7185",
}
ALGO_NAMES = {
    "blank"   : "Blanking",
    "template": "Fixed template",
    "ewma"    : "EWMA template",
    "destd"   : "Dual-ch (DESTD)",
    "gso"     : "GSO",
    "lms"     : "LMS",
    "enlms"   : "ε-NLMS",
    "rls"     : "RLS",
    "ceemdan" : "CEEMDAN",
}

# Metric mode definitions — label, short id, paper reference
METRIC_MODES = [
    ("Current (Inter-stim)",  "inter",    "Your method · 30 ms guard"),
    ("Wang/Chen 2021/23",     "wang",     "SNR=E[e²]/E[(y-e)²] · full signal"),
    ("Sennels 1997",          "sennels",  "MRI_y + PR · adaptive filter paper"),
    ("ASR · Liu/Limnuson",    "asr",      "20·log10(pp_art/pp_residual)"),
    ("Mandrile 2003",         "mandrile", "ARV_norm = |art|/|mwave| ×100%"),
    ("Limnuson RMS reduc.",   "limnuson", "std(art)/std(residual) · RMS ratio"),
    ("Andrews 2023 FFT%",     "andrews",  "% FFT power preserved after filtering"),
]

METRIC_DESCRIPTIONS = {
    "inter": (
        "Inter-stim segments only · 30 ms post-pulse guard excluded\n"
        "SNR = 10·log10(E[clean²] / E[(out-clean)²])\n"
        "SDR = same formula · SDR≈SNR under this definition\n"
        "RMSE in µV · Pearson r · Data loss % (blanking only)\n"
        "Ref: your §2.10 unified benchmarking framework"
    ),
    "wang": (
        "Wang et al. 2021 (GS-APEF/LMS) · Chen et al. 2023 (GSO)\n"
        "SNR  = 10·log10( Σe² / Σ(y-e)² )   [full signal]\n"
        "NRMSE = √mean((y-e)²) / std(e)\n"
        "Pearson r between clean and output\n"
        "Applied over FULL signal (no inter-stim masking)"
    ),
    "sennels": (
        "Sennels et al. 1997 · LMS/RLS adaptive filter paper\n"
        "MRI_y = 10·log10( Σv² / Σy² )\n"
        "  v = volitional EMG (clean), y = filter output\n"
        "PR   = 10·log10( Σx² / Σy² )\n"
        "  x = contaminated input, y = filter output\n"
        "Target: MRI_y near 0 dB · PR as high as possible"
    ),
    "asr": (
        "Liu et al. 2025 (CMOS FPGA) · Limnuson et al. 2014\n"
        "ASR = 20·log10( (V_SA)_pp / (V_SA - V_SAT)_pp )\n"
        "  pp_art = peak-to-peak of original artefact signal\n"
        "  pp_res = peak-to-peak of residual after suppression\n"
        "Target: >20 dB ASR (Liu achieved 20.2 dB in saline)\n"
        "Measures artefact waveform cancellation quality"
    ),
    "mandrile": (
        "Mandrile et al. 2003 · Blanking / spatial characterisation\n"
        "ARV_art  = mean(|art_segment|) in µV\n"
        "ARV_mwave = mean(|mwave_segment|) in µV\n"
        "ARV_norm  = ARV_art / ARV_mwave × 100%\n"
        "Lower ARV_norm = better suppression relative to M-wave\n"
        "Reported range: 46–91% depending on electrode distance"
    ),
    "limnuson": (
        "Limnuson et al. 2014 · IIR template subtraction on FPGA\n"
        "RMS_reduction = std(art) / std(residual)\n"
        "  art      = original artefact signal\n"
        "  residual = output - clean  (what the algorithm left behind)\n"
        "Higher ratio = better artefact suppression\n"
        "Reported: 17× for Aplysia · 5.3× for rat cortex\n"
        "Also reports: recovery latency (ms from spike to neural signal)"
    ),
    "andrews": (
        "Andrews et al. 2023 · AA-IF vs EMD-BF · CEEMDAN paper\n"
        "FFT_pct = sum|FFT(output)|² / sum|FFT(input_clean)|² × 100%\n"
        "  Measures how much of the original spectral content survives\n"
        "  after artefact removal — without being over-filtered\n"
        "Higher % = more signal preserved (less over-suppression)\n"
        "Reported: AA-IF 96±5% · EMD-BF 75±6%  (p<0.001)\n"
        "Companion: EMGrms change before/after as absolute check"
    ),
}

matplotlib.rcParams.update({
    "figure.facecolor" : BG,
    "axes.facecolor"   : SURFACE,
    "axes.edgecolor"   : BORDER,
    "axes.labelcolor"  : MUTED,
    "axes.titlecolor"  : TEXT,
    "axes.grid"        : True,
    "grid.color"       : BORDER,
    "grid.linewidth"   : 0.5,
    "xtick.color"      : MUTED,
    "ytick.color"      : MUTED,
    "xtick.labelsize"  : 7,
    "ytick.labelsize"  : 7,
    "axes.titlesize"   : 9,
    "axes.labelsize"   : 8,
    "legend.fontsize"  : 7.5,
    "legend.framealpha": 0.25,
    "legend.edgecolor" : BORDER,
    "font.family"      : "monospace",
    "text.color"       : TEXT,
})

# ─── explainers ───────────────────────────────────────────────────────────────
EXPLAINERS = {
    "blank": {
        "title": "Blanking",
        "ref"  : "§2.3  ·  Mandrile 2003  ·  Huang 2023",
        "eq"   : "D  =  Tb × fs",
        "lines": [
            "Amplifier input gated off for a fixed window after every pulse.",
            "Signal is zeroed — no recovery attempted.",
            "",
            "  · Zero compute — instant hardware protection",
            "  · Hard data loss scales linearly with stim frequency",
            "  · 50 Hz + 5 ms  →  25% data gone",
            "  · 100 Hz + 10 ms → 100% data loss  (Table 2.1)",
            "",
            "Mandrile 2003: ARV_norm 46–91% depending on electrode",
            "distance. Blanking window validated at 3 ms (6 samples).",
            "Huang 2023: SSAB recovery time 230 µs vs 1.1 s for reset.",
        ],
    },
    "template": {
        "title": "Fixed template subtraction",
        "ref"  : "§2.6.1  ·  Limnuson 2014",
        "eq"   : "y(t)  =  x(t)  -  s_avg(t)",
        "lines": [
            "Constructs an artefact estimate by cycle-locked averaging",
            "of multiple stimulus-locked recordings, then subtracts it.",
            "",
            "  · First n_avg cycles used as causal calibration window",
            "  · Limnuson 2014 (FIR): requires ≥N memory rows",
            "  · Fast once template is built",
            "  · Fails when M-wave recruitment or impedance shifts",
            "",
            "Liu 2025 CMOS: 20.2 dB ASR in saline with impedance",
            "model template (Randles model, 0.75 ms latency).",
        ],
    },
    "ewma": {
        "title": "Adaptive EWMA template",
        "ref"  : "§2.6.2  ·  Limnuson 2014  ·  Eq. 6",
        "eq"   : "s_{k+1}(t) = (1-alpha)*s_k(t) + alpha*x_k(t)",
        "lines": [
            "IIR implementation of template subtraction (Limnuson 2014).",
            "Equivalent to EWMA — single memory row, 3 additions/sample.",
            "",
            "alpha = 1/16 or 1/32 (shift-right operation — no multiplier)",
            "  · Large alpha -> tracks fast, noisier template",
            "  · Small alpha -> stable template, slow to adapt",
            "",
            "  · 17x RMS reduction in Aplysia recordings",
            "  · 5x reduction in rat cortex",
            "  · Spike detection within ~0.5 ms post-stimulus",
            "  · Warm-started from first captured artefact cycle",
        ],
    },
    "destd": {
        "title": "Dual-channel subtraction (DESTD)",
        "ref"  : "§2.5.1  ·  Mandrile 2003  ·  Chen 2023",
        "eq"   : "v(t) = [x1(t)-x2(t) - (x1(t+T)-x2(t+T))] / 2",
        "lines": [
            "Exploits spatial similarity: artefacts appear nearly identically",
            "across adjacent electrodes (~0.9 correlation, Chen 2023).",
            "",
            "Mandrile 2003: normalised ARV independent of waveform shape,",
            "IED, or spatial filter. Electrode distance is the key factor.",
            "",
            "  · No iteration or convergence — pure arithmetic",
            "  · No data lost during suppression",
            "  · Typically 12-16 dB artefact suppression",
            "  · Requires a second recording channel",
        ],
    },
    "gso": {
        "title": "Gram-Schmidt Orthogonalisation (GSO)",
        "ref"  : "§2.5.2  ·  Chen 2023  ·  Eq. 5",
        "eq"   : "z = x - sum_k <x,q_k> q_k",
        "lines": [
            "Chen 2023 (G-S-G): 6th-order GS algorithm on FPGA.",
            "SNR range: −15.26 ± 3.87 dB to −46.19 ± 6.53 dB",
            "NRMSE: 2.28 ± 0.36 to 11.35 ± 2.22 across r=4–20.",
            "",
            "GS-APEF vs LMS-AF: no significant difference in SNR/NRMSE",
            "(two-way ANOVA, p>0.05). Both outperform comb filter.",
            "",
            "  · 13-16 dB SAR improvement reported",
            "  · Single-pass per cycle — no iterative convergence",
            "  · Performance tied to reference channel quality",
        ],
    },
    "lms": {
        "title": "LMS adaptive filter",
        "ref"  : "§2.7.1  ·  Thesis Eq.7  ·  Wang 2021",
        "eq"   : "w[n+1] = w[n] + mu * e[n] * x[n]",
        "lines": [
            "Sample-wise LMS — Thesis Eq.7.",
            "Reference x[n]: signal delayed by one full stim period N.",
            "This aligns the reference with the previous M-wave cycle,",
            "giving the filter a real artefact-shaped reference to cancel.",
            "",
            "  e[n] = ch1[n] - w'x[n]   (error = what's left after cancel)",
            "  w updated each sample — converges over ~100s of samples",
            "",
            "  · mu too large → diverges (weights blow up)",
            "  · mu too small → slow convergence, poor suppression",
            "  · Try mu = 0.01–0.05 for typical FES-EMG amplitudes",
            "  · Wang 2021: LMS-AF vs GS-APEF — no significant difference",
            "    in SNR/NRMSE (p > 0.05, two-way ANOVA)",
        ],
    },
    "enlms": {
        "title": "ε-Normalised LMS",
        "ref"  : "§2.7.1  ·  Thesis Eq.8  ·  Boyer 2023",
        "eq"   : "w[n+1] = w[n] + mu/(‖x‖²+ε) * e[n] * x[n]",
        "lines": [
            "ε-NLMS — Thesis Eq.8. Same one-period-delayed reference",
            "as LMS, but step size normalised by input power.",
            "",
            "  norm = x[n]'x[n] + ε",
            "  w[n+1] = w[n] + (mu/norm) * e[n] * x[n]",
            "",
            "  · Prevents divergence when M-wave amplitude varies",
            "  · ε avoids division by zero during silent periods",
            "  · More stable than LMS for non-stationary recordings",
            "  · Boyer 2023: preferred over LMS for high amplitude var.",
            "  · Same O(order) compute cost as LMS",
        ],
    },
    "rls": {
        "title": "RLS adaptive filter",
        "ref"  : "§2.7.1  ·  Thesis Eqs.9-11",
        "eq"   : "k[n] = P[n-1]x / (λ + x'P[n-1]x)",
        "lines": [
            "RLS — Thesis Eqs.9-11. Minimises TOTAL squared error",
            "over all past samples (not just current like LMS).",
            "Same one-period-delayed reference as LMS/ε-NLMS.",
            "",
            "  k[n] = P[n-1]x / (λ + x'P[n-1]x)    Eq.9",
            "  w[n] = w[n-1] + k[n]*e[n]             Eq.10",
            "  P[n] = (P[n-1] - k[n]x'P[n-1]) / λ   Eq.11",
            "",
            "  · Converges in just a few samples vs ~100s for LMS",
            "  · λ controls forgetting: 0.999 = ~1000 sample memory",
            "  · O(order²) per sample — expensive on MCU",
            "  · The compute gap IS a thesis finding (§2.11)",
        ],
    },
    "ceemdan": {
        "title": "CEEMDAN decomposition",
        "ref"  : "§2.7.2  ·  Andrews 2023  ·  Boyer 2023",
        "eq"   : "x(t) = sum_k IMF_k(t)  [artefact IMFs filtered]",
        "lines": [
            "Andrews 2023 (AA-IF vs EMD-BF): AA-IF preserves 96±5% of",
            "FFT content vs 75±6% for EMD-BF (p<0.001).",
            "",
            "Boyer 2023: CEEMD-IT outperforms SWT, EMD, EMD-IT for",
            "most tested SNR levels while retaining more information.",
            "",
            "  · Strong for nonlinear, amplitude-varying artefacts",
            "  · Handles artefacts distributed over harmonics",
            "  · High compute cost — NOT suitable for real-time MCU",
            "  · The compute overhead IS a documented finding (§2.11)",
        ],
    },
}


# =============================================================================
#  DSP UTILITIES  (unchanged from v2)
# =============================================================================

def gen_emg(fs, duration, amp):
    n = int(fs * duration)
    raw = np.random.randn(n) * amp
    b, a = sp.butter(4, [20, 250], btype="bandpass", fs=fs)
    return sp.filtfilt(b, a, raw)


def _mwave(fs, amp_emg, alpha=220, tau_mw=20e-3):
    dur_s  = 0.06                            # M-wave window duration: 60 ms
    n_mw   = int(fs * dur_s)                 # number of samples in that window
    t      = np.arange(n_mw) / fs            # time axis in seconds

    onset  = int(0.005 * fs)                 # 5 ms onset delay (nerve conduction time before muscle fires)
    f_mw   = 120.0                           # oscillation frequency of the M-wave (Hz)
    mwave  = np.zeros(n_mw)                  # initialise output array to zero

    t_rel  = t[onset:]                       # time axis starting from onset (so decay starts at t=0)
    mwave[onset:] = (
        alpha * t_rel                        # linear ramp — shapes the attack of the wave
        * np.exp(-t_rel / tau_mw)            # exponential decay with time constant tau_mw (20 ms)
        * np.sin(2 * np.pi * f_mw * t_rel)  # 120 Hz sinusoidal oscillation (the compound action potential)
    )

    mw_scale = amp_emg * 80.0               # M-wave is ~80x the background EMG amplitude
    mwave   *= mw_scale / (np.max(np.abs(mwave)) + 1e-30)  # normalise to that target amplitude
                                             # (+1e-30 prevents divide-by-zero)
    mwave   *= 1.0 + np.random.randn() * 0.08  # add ±8% random amplitude jitter per pulse

    return mwave

def gen_artefact(fs, n, stim_freq, amp, emg_amp=300e-6):
    art      = np.zeros(n)                      # output signal, all zeros to start
    interval = max(1, int(fs / stim_freq))      # samples between each stimulus pulse
                                                # e.g. fs=2000, stim_freq=20Hz → interval=100 samples
    pw       = max(2, int(0.0005 * fs))         # pulse width in samples (0.5 ms minimum 2 samples)
    tau_dec  = max(1, int(0.005 * fs))          # exponential decay time constant in samples (5 ms)
    times    = list(range(0, n, interval))      # list of sample indices where each stimulus fires

    mw_template = _mwave(fs, emg_amp)           # pre-compute one M-wave template (reused each pulse)
    mw_len      = len(mw_template)              # length of that template in samples

    t_full    = np.arange(n) / fs              # time axis in seconds for the whole signal
    drift_amp = amp * 0.02                      # drift amplitude = 2% of stimulus amplitude

    drift     = (
        drift_amp * np.sin(2 * np.pi * 0.8 * t_full)           # slow sinusoidal sway at 0.8 Hz
        + drift_amp * 0.5 * np.cumsum(np.random.randn(n))       # random walk (cumulative sum of noise)
                        / np.sqrt(n)                            # normalised so it doesn't grow too large
    )
    art += drift                                # add baseline drift to the whole signal upfront

    for st in times:                            # loop over every stimulus pulse
        pulse_scale = 1.0 + np.random.randn() * 0.05   # ±5% random amplitude jitter per pulse

        half = pw // 2                          # first half of pulse width

        # --- positive phase of biphasic spike ---
        for j in range(half):
            if st + j < n:                      # bounds check
                art[st + j] += amp * pulse_scale   # add positive deflection

        # --- negative phase of biphasic spike ---
        for j in range(half, pw):
            if st + j < n:                      # bounds check
                art[st + j] -= amp * pulse_scale   # subtract for negative deflection

        # --- exponential tail decay after the spike ---
        for j in range(tau_dec * 6):            # run for 6 time constants (≈99% of decay)
            idx = st + pw + j                   # index starts right after the pulse ends
            if idx < n:                         # bounds check
                art[idx] += amp * 0.3 * pulse_scale * np.exp(-j / tau_dec)
                #           ↑ 30% of spike amp  ↑ decays exponentially with tau_dec

        # --- add M-wave starting at the stimulus time ---
        mw_start = st                           # M-wave begins at the stimulus sample
        mw_end   = min(mw_start + mw_len, n)   # clip to signal length
        seg_len  = mw_end - mw_start            # how many samples we can actually write
        if seg_len > 0:
            art[mw_start:mw_end] += mw_template[:seg_len]  # overlay M-wave template

    return art, times, interval
    # returns: the composite artefact signal, the stimulus onset times, and the inter-stimulus interval

def make_ch2(emg, art, fs, emg_amp=300e-6, noise=5e-6):
    emg2    = gen_emg(fs, len(emg) / fs, np.std(emg))
    # EMG correlation ~0.15 — mostly independent across channels.
    # This is correct for spatial methods (GSO, DESTD) which exploit
    # the EMG difference between channels to isolate the artefact.
    # For adaptive filters (LMS/RLS) the reference needs similar artefact
    # but different EMG — same requirement, so 15% correlation works for both.
    # The convergence problem in LMS/RLS is fixed via signal normalisation,
    # not by changing channel correlation.
    emg_ch2 = emg * 0.15 + emg2 * 0.85
    art_ch2 = art * (0.92 + np.random.randn() * 0.02)
    return emg_ch2 + art_ch2 + np.random.randn(len(emg)) * noise


def load_emg_file(path, fs_app):
    """
    Load EMG file. For Ninapro .mat files with a restimulus column:
      - Rows where restimulus != 0  ->  active EMG  ->  used as clean signal
      - Rows where restimulus == 0  ->  rest noise   ->  stored as dynamic noise

    Returns: data, actual_fs, info, ch2_data, noise_ch1, noise_ch2
    """
    ext       = os.path.splitext(path)[1].lower()
    noise_ch1 = None
    noise_ch2 = None
    try:
        actual_fs = fs_app
        ch_info   = ""
        ch2_data  = None
        if ext == ".mat":
            from scipy.io import loadmat
            mat = loadmat(path, squeeze_me=True)
            # Ninapro-style: has restimulus label column
            if "restimulus" in mat and "emg" in mat:
                emg_all    = np.array(mat["emg"], dtype=float)
                restimulus = np.array(mat["restimulus"]).flatten()
                actual_fs  = 2000
                if emg_all.ndim == 1:
                    emg_all = emg_all[:, np.newaxis]
                n_ch        = emg_all.shape[1]
                active_mask = restimulus != 0
                rest_mask   = restimulus == 0
                # Active rows -> clean EMG signal
                if active_mask.sum() > actual_fs * 2:
                    active_emg = emg_all[active_mask, :]
                    ch_info    = (f"{n_ch} ch Ninapro  |  "
                                  f"active: {active_mask.sum()/actual_fs:.0f}s")
                else:
                    active_emg = emg_all
                    ch_info    = f"{n_ch} ch Ninapro [no active mask]"
                data     = active_emg[:, 0].copy()
                ch2_data = active_emg[:, 1].copy() if n_ch > 1 else None
                data -= data.mean()
                if ch2_data is not None:
                    ch2_data -= ch2_data.mean()
                # Rest rows -> dynamic noise (unit std; _get_noise_segment scales to 5 uV)
                if rest_mask.sum() > actual_fs * 2:
                    rest_emg  = emg_all[rest_mask, :]
                    noise_ch1 = rest_emg[:, 0].copy()
                    noise_ch1 -= noise_ch1.mean()
                    s1 = noise_ch1.std()
                    if s1 > 1e-12:
                        noise_ch1 /= s1
                    if n_ch > 1:
                        noise_ch2 = rest_emg[:, 1].copy()
                        noise_ch2 -= noise_ch2.mean()
                        s2 = noise_ch2.std()
                        if s2 > 1e-12:
                            noise_ch2 /= s2
                    ch_info += f"  |  noise: {rest_mask.sum()/actual_fs:.0f}s rest"
            # Generic .mat (no restimulus)
            elif "emg" in mat:
                emg_all = np.array(mat["emg"], dtype=float)
                if emg_all.ndim == 1:
                    data = emg_all; ch2_data = None
                    ch_info = "1 channel"
                else:
                    data = emg_all[:, 0]
                    ch2_data = emg_all[:, 1] if emg_all.shape[1] > 1 else None
                    ch_info = f"{emg_all.shape[1]} ch — using ch0"
                actual_fs = 2000
            else:
                found = False; ch2_data = None
                for k, v in mat.items():
                    if k.startswith("__"): continue
                    arr = np.array(v, dtype=float).flatten()
                    if arr.size > 100:
                        data = arr; ch_info = f"key='{k}'"; found = True; break
                if not found:
                    raise ValueError("No suitable array found in .mat file.")
        elif ext == ".npy":
            raw  = np.load(path).astype(float)
            data = raw[:, 0] if raw.ndim == 2 else raw.flatten()
            ch_info = "npy"
        elif ext in (".csv", ".txt"):
            raw  = np.loadtxt(path, delimiter=",")
            data = (raw[:, -1] if raw.ndim == 2 else raw).astype(float)
            ch_info = "csv/txt"
        else:
            raise ValueError(f"Unsupported type: {ext}. Use .mat .npy .csv .txt")
        std = np.std(data)
        if std > 1e-9:
            data = data / std * 300e-6
            if ch2_data is not None:
                std2 = np.std(ch2_data)
                if std2 > 1e-9:
                    ch2_data = ch2_data / std2 * 300e-6
                else:
                    ch2_data = None
        else:
            raise ValueError("Signal appears to be all zeros.")
        info = f"{os.path.basename(path)}  ·  {len(data)/actual_fs:.1f} s  ·  {ch_info}"
        return data, actual_fs, info, ch2_data, noise_ch1, noise_ch2
    except Exception as e:
        raise RuntimeError(str(e))


# =============================================================================
#  METRICS  — all 5 modes
# =============================================================================

def _make_inter_stim_mask(n, times, fs, stim_freq, guard_ms=30):
    guard = int((guard_ms / 1000.0) * fs)
    mask  = np.ones(n, dtype=bool)
    for st in times:
        end = min(st + guard, n)
        mask[st:end] = False
    return mask


def compute_metrics(clean, rec, algo, blank_ms, stim_freq,
                    times=None, fs=4000, guard_ms=30,
                    mode="inter",
                    art=None):
    """
    Unified metric computation supporting 5 modes:

    inter    : inter-stim segments, 30 ms guard  (your original)
    wang     : Wang/Chen full-signal SNR + NRMSE
    sennels  : MRI_y + PR (Sennels 1997 adaptive filter paper)
    asr      : Peak-to-peak ASR (Liu/Limnuson FPGA template papers)
    mandrile : ARV_norm = |art| / |mwave| ×100%  (Mandrile 2003)
    """
    c = clean.copy()
    r = rec.copy()
    eps = 1e-30

    # ── Mode: inter-stim (original) ──────────────────────────────────────────
    if mode == "inter":
        if times is not None and len(times) > 0:
            mask = _make_inter_stim_mask(len(c), times, fs, stim_freq, guard_ms)
            if mask.sum() > 64:
                c = c[mask]; r = r[mask]
        s_p  = np.mean(c**2) + eps
        n_p  = np.mean((r - c)**2) + eps
        snr  = 10 * np.log10(s_p / n_p)
        sdr  = snr
        rmse = float(np.sqrt(np.mean((r - c)**2)) * 1e6)
        corr = float(np.corrcoef(c, r)[0, 1])
        loss = (blank_ms / 1000) * stim_freq * 100 if algo == "blank" else 0.0
        return dict(
            mode="inter",
            snr=snr, sdr=sdr, rmse=rmse, r=corr, loss=loss,
            # extra fields blank for this mode
            nrmse=rmse / (np.std(c) * 1e6 + eps),
            mri_y=0.0, pr=0.0, asr=0.0, arv_norm=0.0, rms_red=0.0, fft_pct=0.0,
        )

    # ── Mode: Wang/Chen (full-signal, clean-referenced) ──────────────────────
    elif mode == "wang":
        # SNR = 10*log10( sum(e^2) / sum((y-e)^2) )
        # where e = clean EMG, y = output
        num  = np.sum(c**2) + eps
        den  = np.sum((r - c)**2) + eps
        snr  = 10 * np.log10(num / den)
        rmse = float(np.sqrt(np.mean((r - c)**2)) * 1e6)
        nrmse = float(np.sqrt(np.mean((r - c)**2)) / (np.std(c) + eps))
        corr  = float(np.corrcoef(c, r)[0, 1])
        loss  = (blank_ms / 1000) * stim_freq * 100 if algo == "blank" else 0.0
        return dict(
            mode="wang",
            snr=snr, sdr=snr, rmse=rmse, r=corr, loss=loss,
            nrmse=nrmse, mri_y=0.0, pr=0.0, asr=0.0, arv_norm=0.0, rms_red=0.0, fft_pct=0.0,
        )

    # ── Mode: Sennels 1997 (MRI_y + PR) ──────────────────────────────────────
    elif mode == "sennels":
        # MRI_y = 10*log10( sum(v^2) / sum(y^2) )
        # v = volitional EMG (clean), y = output
        # PR = 10*log10( sum(x^2) / sum(y^2) )
        # x = contaminated input
        x = clean + (art if art is not None else np.zeros_like(clean))
        v = clean
        y = rec
        mri_y = 10 * np.log10((np.sum(v**2) + eps) / (np.sum(y**2) + eps))
        pr    = 10 * np.log10((np.sum(x**2) + eps) / (np.sum(y**2) + eps))
        rmse  = float(np.sqrt(np.mean((y - v)**2)) * 1e6)
        corr  = float(np.corrcoef(v, y)[0, 1])
        loss  = (blank_ms / 1000) * stim_freq * 100 if algo == "blank" else 0.0
        # SNR slot reused for MRI_y, SDR for PR so strip fits
        return dict(
            mode="sennels",
            snr=mri_y, sdr=pr, rmse=rmse, r=corr, loss=loss,
            nrmse=0.0, mri_y=mri_y, pr=pr, asr=0.0, arv_norm=0.0,
            rms_red=0.0, fft_pct=0.0,
        )

    # ── Mode: ASR peak-to-peak (Liu 2025 / Limnuson 2014) ────────────────────
    elif mode == "asr":
        # ASR = 20*log10( pp_original_art / pp_residual )
        if art is not None:
            pp_art = float(np.ptp(art))
        else:
            # estimate artefact from contaminated - clean
            ch1 = clean + np.zeros_like(clean)  # fallback
            pp_art = float(np.ptp(rec - clean)) * 10  # crude estimate

        residual = rec - clean
        pp_res   = float(np.ptp(residual)) + eps
        pp_art   = max(float(np.ptp(art)) if art is not None else pp_res * 10, eps)
        asr_val  = 20 * np.log10(pp_art / pp_res)

        rmse = float(np.sqrt(np.mean((rec - clean)**2)) * 1e6)
        corr = float(np.corrcoef(clean, rec)[0, 1])
        loss = (blank_ms / 1000) * stim_freq * 100 if algo == "blank" else 0.0
        return dict(
            mode="asr",
            snr=asr_val, sdr=asr_val, rmse=rmse, r=corr, loss=loss,
            nrmse=0.0, mri_y=0.0, pr=0.0, asr=asr_val, arv_norm=0.0,
            rms_red=0.0, fft_pct=0.0,
        )

    # ── Mode: Mandrile 2003 (ARV_norm) ───────────────────────────────────────
    elif mode == "mandrile":
        # ARV_norm = mean|art_segment| / mean|mwave_segment| × 100%
        # art_segment = first 3 ms post-pulse (Mandrile: 6 samples @ 2048 Hz)
        # mwave_segment = remainder of inter-stimulus window
        art_seg_ms   = 3    # ms for artefact window (Mandrile 2003)
        art_samples  = max(1, int((art_seg_ms / 1000) * fs))

        arv_arts  = []
        arv_mwaves = []

        interval = max(1, int(fs / stim_freq))
        # Use the artefact signal for the artefact window measurement.
        # This avoids the blanking edge case where rec is zeroed and
        # gives mean|art_seg|=0, making ARV_norm=0 (physically meaningless).
        # Mandrile measures residual artefact in the output, so for blanking
        # the residual IS zero — but we still want a meaningful denominator.
        art_ref = art if art is not None else np.zeros_like(rec)
        for st in (times if times else []):
            # artefact window: first art_samples after pulse
            art_end = min(st + art_samples, len(rec))
            if art_end <= st:
                continue
            # residual artefact = what's left in output after suppression
            residual_seg = rec[st:art_end] - clean[st:art_end] if clean is not None else rec[st:art_end]
            arv_arts.append(np.mean(np.abs(residual_seg)))

            # mwave window: art_samples to end of period (use output)
            mw_start = art_end
            mw_end   = min(st + interval, len(rec))
            if mw_end > mw_start:
                arv_mwaves.append(np.mean(np.abs(rec[mw_start:mw_end])))

        if arv_arts and arv_mwaves:
            arv_norm = (np.mean(arv_arts) /
                        (np.mean(arv_mwaves) + eps)) * 100.0
        else:
            arv_norm = 0.0

        rmse = float(np.sqrt(np.mean((rec - clean)**2)) * 1e6)
        corr = float(np.corrcoef(clean, rec)[0, 1])
        loss = (blank_ms / 1000) * stim_freq * 100 if algo == "blank" else 0.0
        return dict(
            mode="mandrile",
            snr=arv_norm, sdr=0.0, rmse=rmse, r=corr, loss=loss,
            nrmse=0.0, mri_y=0.0, pr=0.0, asr=0.0, arv_norm=arv_norm,
            rms_red=0.0, fft_pct=0.0,
        )

    # ── Mode: Limnuson 2014 (RMS reduction ratio) ─────────────────────────────
    elif mode == "limnuson":
        # RMS_reduction = std(artefact) / std(output - clean)
        # Limnuson: 17x (Aplysia), 5.3x (rat). Higher = better.
        art_signal = art if art is not None else np.zeros_like(clean)
        residual   = rec - clean
        rms_red    = float(np.std(art_signal) / (np.std(residual) + eps))

        # Liu 2014 companion: CC + RMSE on M-wave window (5–50 ms post-pulse)
        mw_s = int(0.005 * fs); mw_e = int(0.050 * fs)
        mw_recs = []; mw_clns = []
        for st in (times if times else []):
            s = st + mw_s; e = min(st + mw_e, len(rec))
            if e > s:
                mw_recs.append(rec[s:e]); mw_clns.append(clean[s:e])
        if mw_recs:
            mr = np.concatenate(mw_recs); mc = np.concatenate(mw_clns)
            mw_cc  = float(np.corrcoef(mc, mr)[0, 1])
            mw_rms = float(np.sqrt(np.mean((mr - mc)**2)) * 1e6)
        else:
            mw_cc = 0.0; mw_rms = 0.0

        loss = (blank_ms / 1000) * stim_freq * 100 if algo == "blank" else 0.0
        return dict(
            mode="limnuson",
            snr=rms_red, sdr=mw_cc, rmse=mw_rms, r=mw_cc, loss=loss,
            nrmse=0.0, mri_y=0.0, pr=0.0, asr=0.0, arv_norm=0.0,
            rms_red=rms_red, fft_pct=0.0,
        )

    # ── Mode: Andrews 2023 (FFT spectral preservation %) ─────────────────────
    elif mode == "andrews":
        # FFT_pct vs input  = sum|FFT(out)|² / sum|FFT(contaminated)|² × 100%
        # FFT_pct vs clean  = sum|FFT(out)|² / sum|FFT(clean)|²        × 100%
        # Andrews: AA-IF 96±5% · EMD-BF 75±6%  (vs contaminated input)
        ch1_sig = clean + (art if art is not None else np.zeros_like(clean))
        psd_out   = np.sum(np.abs(np.fft.rfft(rec))**2)
        psd_in    = np.sum(np.abs(np.fft.rfft(ch1_sig))**2) + eps
        psd_clean = np.sum(np.abs(np.fft.rfft(clean))**2)   + eps
        fft_vs_in    = float(psd_out / psd_in    * 100.0)
        # FFT% vs clean: cap at 200% to avoid huge numbers when output
        # still contains artefact power (blanking, template methods).
        # Andrews reports this companion metric only for CEEMDAN/AA-IF
        # where the output is already close to clean. For other methods
        # the raw ratio is misleading — report EMG RMS ratio instead.
        emg_rms_ratio = float(np.std(rec) / (np.std(clean) + eps))
        # fft_vs_clean: ratio of output PSD to clean PSD, capped at 200%
        fft_vs_clean = float(min(psd_out / psd_clean * 100.0, 200.0))

        rmse = float(np.sqrt(np.mean((rec - clean)**2)) * 1e6)
        corr = float(np.corrcoef(clean, rec)[0, 1])
        loss = (blank_ms / 1000) * stim_freq * 100 if algo == "blank" else 0.0
        return dict(
            mode="andrews",
            snr=fft_vs_in, sdr=fft_vs_clean,
            rmse=rmse, r=corr, loss=loss,
            nrmse=emg_rms_ratio, mri_y=0.0, pr=0.0, asr=0.0, arv_norm=0.0,
            rms_red=0.0, fft_pct=fft_vs_in,
        )

    else:
        raise ValueError(f"Unknown metric mode: {mode}")


def _limnuson_rms_reduction(art, rec, clean, eps=1e-30):
    """
    Limnuson 2014: RMS reduction factor.
    std(original artefact) / std(residual after suppression)
    Reported as a plain ratio (e.g. 17× for Aplysia, 5.3× for rat).
    Also returns latency proxy — first sample where residual < threshold.
    """
    residual = rec - clean
    rms_art  = float(np.std(art) + eps)
    rms_res  = float(np.std(residual) + eps)
    return rms_art / rms_res


def _andrews_fft_pct(rec, ch1, eps=1e-30):
    """
    Andrews 2023: percentage of FFT spectral power preserved.
    sum|FFT(output)|² / sum|FFT(contaminated_input)|² × 100%

    Interpretation:
      ~100% = algorithm preserves all spectral content (may not remove artefact)
      ~75%  = EMD-BF result (over-filters, removes too much)
      ~96%  = AA-IF result (targeted removal, preserves EMG spectrum)
    The ideal is close to what the CLEAN signal would give, not 100% of
    the contaminated signal (which includes artefact power).
    """
    psd_out = np.sum(np.abs(np.fft.rfft(rec))**2)
    psd_in  = np.sum(np.abs(np.fft.rfft(ch1))**2)
    return float(psd_out / (psd_in + eps) * 100.0)


# =============================================================================
#  METRIC STRIP LABEL MAPPING  — changes labels depending on mode
# =============================================================================

STRIP_LABELS = {
    "inter"   : [("algo","ALGORITHM",150),("snr","SNR",110),("sdr","SDR",110),
                 ("rmse","RMSE (µV)",120),("r","CORR r",110),
                 ("loss","DATA LOSS",105),("lat","LATENCY",130)],
    "wang"    : [("algo","ALGORITHM",150),("snr","SNR (dB)",110),("sdr","—",110),
                 ("rmse","RMSE (µV)",120),("r","CORR r",110),
                 ("loss","DATA LOSS",105),("lat","LATENCY",130)],
    "sennels" : [("algo","ALGORITHM",150),("snr","MRI_y (dB)",120),("sdr","PR (dB)",110),
                 ("rmse","RMSE (µV)",120),("r","CORR r",110),
                 ("loss","DATA LOSS",105),("lat","LATENCY",130)],
    "asr"     : [("algo","ALGORITHM",150),("snr","ASR (dB)",120),("sdr","—",80),
                 ("rmse","RMSE (µV)",120),("r","CORR r",110),
                 ("loss","DATA LOSS",105),("lat","LATENCY",130)],
    "mandrile": [("algo","ALGORITHM",150),("snr","ARV_norm%",120),("sdr","—",80),
                 ("rmse","RMSE (µV)",120),("r","CORR r",110),
                 ("loss","DATA LOSS",105),("lat","LATENCY",130)],
    "limnuson": [("algo","ALGORITHM",150),("snr","RMS reduc×",130),("sdr","Mwave CC",110),
                 ("rmse","Mwave RMS",120),("r","CORR r",110),
                 ("loss","DATA LOSS",105),("lat","LATENCY",130)],
    "andrews" : [("algo","ALGORITHM",150),("snr","FFT% (in)",120),("sdr","FFT% (cln)",120),
                 ("rmse","RMSE (µV)",110),("r","CORR r",110),
                 ("loss","DATA LOSS",105),("lat","LATENCY",130)],
}

STRIP_FORMATS = {
    "inter"   : {"snr":"{:+.1f} dB","sdr":"{:+.1f} dB","rmse":"{:.2f} µV",
                 "r":"{:.4f}","loss":"{:.1f}%"},
    "wang"    : {"snr":"{:+.1f} dB","sdr":"—","rmse":"{:.2f} µV",
                 "r":"{:.4f}","loss":"{:.1f}%"},
    "sennels" : {"snr":"{:+.1f} dB","sdr":"{:+.1f} dB","rmse":"{:.2f} µV",
                 "r":"{:.4f}","loss":"{:.1f}%"},
    "asr"     : {"snr":"{:.1f} dB","sdr":"—","rmse":"{:.2f} µV",
                 "r":"{:.4f}","loss":"{:.1f}%"},
    "mandrile": {"snr":"{:.1f}%","sdr":"—","rmse":"{:.2f} µV",
                 "r":"{:.4f}","loss":"{:.1f}%"},
    "limnuson": {"snr":"{:.1f}×","sdr":"{:.4f}","rmse":"{:.2f} µV",
                 "r":"{:.4f}","loss":"{:.1f}%"},
    "andrews" : {"snr":"{:.1f}%","sdr":"{:.1f}%","rmse":"{:.2f} µV",
                 "r":"{:.4f}","loss":"{:.1f}%"},
}


# =============================================================================
#  ALGORITHMS  (unchanged from v2)
# =============================================================================

def _cycle_bounds(times, n, interval=None):
    bounds = []
    if times is None: return bounds
    for i, st in enumerate(times):
        if st >= n: break
        if i + 1 < len(times):
            end = min(times[i + 1], n)
        elif interval is not None:
            end = min(st + interval, n)
        else:
            end = n
        if end > st: bounds.append((st, end))
    return bounds


def _build_gso_basis(ref_seg, order):
    taps = min(order, len(ref_seg))
    if taps <= 0: return None
    X = np.zeros((len(ref_seg), taps))
    for k in range(taps):
        X[k:, k] = ref_seg[:len(ref_seg) - k]
    if not np.any(np.abs(X) > 1e-12): return None
    q, r = np.linalg.qr(X, mode="reduced")
    keep = np.abs(np.diag(r)) > 1e-10
    if not np.any(keep): return None
    return q[:, keep]


def _sennels_frames(ch1, times, interval, n):
    """
    Sennels 1997 §II-B: divide the signal into frames of N samples,
    where N = fs / stim_freq (one stimulation period per frame).

    Returns a list of (start_idx, frame_vector) tuples, one per
    stimulation pulse.  Each frame is exactly `interval` samples long,
    zero-padded at the end if necessary.
    """
    frames = []
    for st in times:
        end = min(st + interval, n)
        seg = ch1[st:end]
        if len(seg) < interval:
            seg = np.pad(seg, (0, interval - len(seg)))
        frames.append((st, seg.copy()))
    return frames


def _sennels_block_ls(frames, M, interval):
    """
    Sennels 1997 Eq. (6)/(7): for each frame j (the 'present' frame),
    solve the least-squares problem

        Φ b = Θ   →   b = Φ⁻¹ Θ

    where
        Φ_{r,s} = s_r · s_s   (dot products of previous frames)
        Θ_l     = s_0 · s_l   (dot product of present with each prev)
        s_0 = present frame,   s_1…s_M = M previous frames

    Returns predicted frame  ŷ₀ = s_0 - Σ b_j s_j

    The final output is scaled per Eq. (10):
        ỹ(i) = y(i) / sqrt(1 + Σ b_j²)
    so output power equals volitional EMG power (MRI_y → 0 dB).
    """
    out_frames = []
    for idx in range(len(frames)):
        st0, s0 = frames[idx]

        # collect M previous frames
        prev = []
        for lag in range(1, M + 1):
            if idx - lag >= 0:
                prev.append(frames[idx - lag][1])

        if len(prev) == 0:
            # no history yet — pass through unchanged
            out_frames.append((st0, s0.copy()))
            continue

        actual_M = len(prev)
        # Build Φ (M×M) and Θ (M,) using inner products
        Phi = np.zeros((actual_M, actual_M))
        Theta = np.zeros(actual_M)
        for r in range(actual_M):
            Theta[r] = float(s0 @ prev[r])
            for s in range(actual_M):
                Phi[r, s] = float(prev[r] @ prev[s])

        # Solve  Φ b = Θ  (least squares — handles near-singular Φ)
        try:
            b, _, _, _ = np.linalg.lstsq(Phi, Theta, rcond=None)
        except np.linalg.LinAlgError:
            out_frames.append((st0, s0.copy()))
            continue

        # predicted frame  ŷ₀ = s₀ - Σ bⱼ sⱼ
        predicted_mwave = sum(b[j] * prev[j] for j in range(actual_M))
        y0 = s0 - predicted_mwave

        # Eq. (10): scale so output RMS ≈ volitional EMG RMS
        scale = np.sqrt(1.0 + float(np.sum(b ** 2)))
        y0_scaled = y0 / (scale + 1e-30)

        out_frames.append((st0, y0_scaled))

    return out_frames


def algo_blanking(ch1, times, fs, blank_ms):
    t0 = time.perf_counter(); out = ch1.copy()
    bs = int((blank_ms / 1000) * fs)
    for st in times: out[st: st + bs] = 0.0
    return out, (time.perf_counter() - t0) / len(ch1) * fs * 1e3


def algo_fixed_template(ch1, times, interval, fs, n_avg=10):
    t0 = time.perf_counter(); n = len(ch1); out = ch1.copy()
    bounds = _cycle_bounds(times, n, interval)
    if len(bounds) < 2: return out, 0.0
    n_use = min(max(1, int(n_avg)), len(bounds) - 1)
    template = np.zeros(interval); counts = np.zeros(interval)
    for st, end in bounds[:n_use]:
        seg_len = min(interval, end - st)
        template[:seg_len] += ch1[st:st + seg_len]; counts[:seg_len] += 1
    valid = counts > 0
    if not np.any(valid): return out, (time.perf_counter() - t0) / n * fs * 1e3
    template[valid] /= counts[valid]
    for st, end in bounds[n_use:]:
        seg_len = min(interval, end - st)
        out[st:st + seg_len] -= template[:seg_len]
    return out, (time.perf_counter() - t0) / n * fs * 1e3


def algo_ewma_template(ch1, times, interval, fs, alpha=0.1):
    t0 = time.perf_counter(); n = len(ch1); out = ch1.copy()
    bounds = _cycle_bounds(times, n, interval)
    if len(bounds) < 2: return out, 0.0
    template = np.zeros(interval)
    first_st, first_end = bounds[0]
    first_len = min(interval, first_end - first_st)
    template[:first_len] = ch1[first_st:first_st + first_len]
    for st, end in bounds[1:]:
        seg_len = min(interval, end - st)
        seg = ch1[st:st + seg_len]
        out[st:st + seg_len] -= template[:seg_len]
        template[:seg_len] = (1 - alpha) * template[:seg_len] + alpha * seg
    return out, (time.perf_counter() - t0) / n * fs * 1e3


def algo_destd(ch1, ch2, times, interval, fs):
    t0 = time.perf_counter(); diff = ch1 - ch2; out = diff.copy()
    bounds = _cycle_bounds(times, len(ch1), interval)
    for i in range(1, len(bounds)):
        st, end = bounds[i]; prev_st, prev_end = bounds[i - 1]
        cur = diff[st:end]; prev = diff[prev_st:prev_end]
        seg_len = min(len(cur), len(prev))
        if seg_len > 0:
            out[st:st + seg_len] = (cur[:seg_len] - prev[:seg_len]) / 2.0
        if seg_len < len(cur):
            out[st + seg_len:end] = cur[seg_len:]
    return out, (time.perf_counter() - t0) / len(ch1) * fs * 1e3


def algo_gso(ch1, ref_ch, times, interval, fs, order=6):
    t0 = time.perf_counter(); out = ch1.copy()
    bounds = _cycle_bounds(times, len(ch1), interval)
    for st, end in bounds:
        sig_seg = ch1[st:end]; ref_seg = ref_ch[st:end]
        basis = _build_gso_basis(ref_seg, order)
        if basis is None: continue
        proj = basis @ (basis.T @ sig_seg)
        out[st:end] = sig_seg - proj
    return out, (time.perf_counter() - t0) / len(ch1) * fs * 1e3


def algo_lms(ch1, ch2, times, interval, fs, mu, order=6):
    """
    Thesis §2.7.1 Eq.7 — sample-wise LMS adaptive filter.

        w[n+1] = w[n] + mu * e[n] * x[n]

    Reference: ch2 (second electrode channel) — carries same artefact
    as ch1 but different volitional EMG. Filter learns to cancel the
    shared artefact component. Matches Wang 2021 (LMS-AF).

    Signal normalisation: ch1 and ch2 are scaled to unit variance
    before filtering so mu=0.01 is meaningful regardless of artefact
    amplitude. Output is rescaled back to original units.
    This fixes convergence at high artefact-to-EMG ratios (>10×).
    """
    t0  = time.perf_counter()
    n   = len(ch1)

    # Normalise to unit variance — makes mu amplitude-independent
    scale = float(np.std(ch1)) + 1e-30
    ch1_n = ch1 / scale
    ref_n = ch2 / scale   # same scale so relative amplitudes preserved

    out_n = np.zeros(n)
    w     = np.zeros(order)

    for i in range(order, n):
        x = ref_n[i - order:i][::-1]
        e = ch1_n[i] - w @ x
        w = w + mu * e * x
        out_n[i] = e

    out_n[:order] = ch1_n[:order]

    # Rescale back to original units
    out = out_n * scale

    return out, (time.perf_counter() - t0) / n * fs * 1e3


def algo_enlms(ch1, ch2, times, interval, fs, mu, order=32, eps=1e-6):
    """
    Thesis §2.7.1 Eq.8 — ε-Normalised LMS.

        w[n+1] = w[n] + (mu / (‖x[n]‖² + ε)) * e[n] * x[n]

    Reference: ch2 (second electrode channel).
    Step size normalised by input power — prevents divergence when
    M-wave amplitude varies across cycles (Boyer 2023).

    Signal normalisation applied before filtering for the same reason
    as algo_lms — makes mu and eps meaningful at any artefact amplitude.
    The per-sample ‖x‖² normalisation handles fast amplitude variation;
    the global scale normalisation handles the baseline amplitude offset.
    """
    t0  = time.perf_counter()
    n   = len(ch1)

    # Normalise to unit variance
    scale = float(np.std(ch1)) + 1e-30
    ch1_n = ch1 / scale
    ref_n = ch2 / scale

    out_n = np.zeros(n)
    w     = np.zeros(order)

    for i in range(order, n):
        x    = ref_n[i - order:i][::-1]
        e    = ch1_n[i] - w @ x
        norm = float(x @ x) + eps
        w    = w + (mu / norm) * e * x
        out_n[i] = e

    out_n[:order] = ch1_n[:order]

    # Rescale back to original units
    out = out_n * scale

    return out, (time.perf_counter() - t0) / n * fs * 1e3


def algo_rls(ch1, ch2, times, interval, fs, lam=0.999, order=16):
    """
    Thesis §2.7.1 Eqs.9-11 — Recursive Least Squares.

        k[n]   = P[n-1]x / (λ + x'P[n-1]x)       Eq.9
        w[n]   = w[n-1] + k[n] * e[n]              Eq.10
        P[n]   = (1/λ)(P[n-1] - k[n] x' P[n-1])   Eq.11

    Reference: ch2 (second electrode channel).
    Converges in a few samples vs ~100s for LMS — but O(order²)
    compute cost makes it impractical on MCU-class hardware.
    This is a documented thesis finding (§2.11).

    Signal normalisation applied before filtering (unit variance)
    so lambda and the initial P=I are meaningful at any signal amplitude.

    P matrix reset: P is reinitialised to I every stimulation period
    (interval samples). This prevents the well-known numerical divergence
    of standard RLS over long recordings caused by repeated division by λ.
    Weight vector w is carried across resets so learned cancellation is
    preserved. This is consistent with Sennels 1997 frame-by-frame
    processing and standard practice in embedded RLS implementations
    (Haykin, Adaptive Filter Theory §9.4).
    """
    t0  = time.perf_counter()
    n   = len(ch1)

    # Normalise to unit variance
    scale = float(np.std(ch1)) + 1e-30
    ch1_n = ch1 / scale
    ref_n = ch2 / scale

    out_n    = np.zeros(n)
    w        = np.zeros(order)
    P        = np.eye(order)
    N_reset  = max(interval, order * 4)   # reset P every stim period

    for i in range(order, n):
        # Reset P matrix every N_reset samples to prevent divergence.
        # w is preserved so learned weights carry over.
        if (i - order) % N_reset == 0 and i > order:
            P = np.eye(order)

        x  = ref_n[i - order:i][::-1]
        Px = P @ x
        denom = lam + float(x @ Px)
        if abs(denom) < 1e-10:          # guard against numerical zero
            out_n[i] = ch1_n[i]
            continue
        k  = Px / denom
        e  = ch1_n[i] - float(w @ x)
        w  = w + k * e
        P  = (P - np.outer(k, Px)) / lam
        out_n[i] = e

    out_n[:order] = ch1_n[:order]

    # Rescale back to original units
    out = out_n * scale

    return out, (time.perf_counter() - t0) / n * fs * 1e3


def _emd_sift(sig, max_imfs=5, max_iter=8):
    from scipy.interpolate import CubicSpline
    imfs = []; resid = sig.copy(); n = len(sig); t = np.arange(n)
    for _ in range(max_imfs):
        h = resid.copy(); prev_sd = np.inf
        for _ in range(max_iter):
            maxima = (np.diff(np.sign(np.diff(h))) < 0).nonzero()[0] + 1
            minima = (np.diff(np.sign(np.diff(h))) > 0).nonzero()[0] + 1
            if len(maxima) < 3 or len(minima) < 3: break
            try:
                env_max = CubicSpline(maxima, h[maxima])(t)
                env_min = CubicSpline(minima, h[minima])(t)
            except Exception: break
            mean_env = (env_max + env_min) / 2.0; h_new = h - mean_env
            sd = float(np.sum((h_new - h) ** 2) / (np.sum(h ** 2) + 1e-30))
            h = h_new
            if sd < 0.2 or abs(sd - prev_sd) < 1e-6: break
            prev_sd = sd
        imfs.append(h); resid = resid - h
        if np.std(resid) < 1e-14 * np.std(sig): break
    imfs.append(resid)
    return imfs


def algo_ceemdan(ch1, times, interval, fs, stim_freq, n_imfs=5):
    """
    Andrews 2023 AA-IF (Artifact Adaptive Ideal Filtering) approach.

    The paper works in the FREQUENCY DOMAIN, not purely EMD:
    1. FFT the signal
    2. Identify frequency bins contaminated by stim artefact —
       spikes at stim_freq and its harmonics (up to Nyquist)
    3. For each harmonic spike: measure surrounding FFT power
       as a threshold, blank contaminated bins (ideal notch)
    4. Reconstruct via inverse FFT

    Then apply simplified EMD to handle residual nonlinear components.

    This matches Andrews 2023 Eq.1-7 and their reported 96±5%
    FFT content preservation vs 75±6% for EMD-BF.
    """
    t0 = time.perf_counter()
    n  = len(ch1)

    # ── Step 1: FFT ──────────────────────────────────────────────────
    X     = np.fft.rfft(ch1)
    freqs = np.fft.rfftfreq(n, 1.0 / fs)
    X_out = X.copy()

    # ── Step 2: Identify & notch artefact harmonics ──────────────────
    # Andrews: search for artifact spikes at stim_freq harmonics
    # Cluster width ~2 Hz around each harmonic (Andrews Eq.1)
    # Andrews 2023 AA-IF: notch genuine artefact harmonic spikes.
    # Threshold 2.0× median background: FES harmonic spikes are typically
    # 5-20× above local noise floor, so 2× catches them without eating
    # into EMG sidebands. cluster_bins=2 gives a narrow but non-zero notch.
    # Only notch first 10 harmonics — higher ones are in the noise floor.
    df           = freqs[1] - freqs[0] if len(freqs) > 1 else 1.0
    cluster_bins = 2   # 2 bins either side of harmonic centre

    max_harmonics = 10
    harmonics = []
    k = 1
    while k * stim_freq < fs / 2.0 and k <= max_harmonics:
        harmonics.append(k * stim_freq)
        k += 1

    for h_freq in harmonics:
        h_idx = int(round(h_freq / df))
        if h_idx >= len(freqs):
            break

        # Local background: 20 bins either side, skipping the notch window
        bg_bins = []
        for offset in range(-20, 21):
            bi = h_idx + offset
            if abs(offset) > cluster_bins and 0 <= bi < len(X):
                bg_bins.append(abs(X[bi]))

        if not bg_bins:
            continue

        bg_mean   = float(np.mean(bg_bins))    # Andrews 2023: arithmetic mean
        threshold = bg_mean * 2.0   # 2× mean — genuine spike threshold

        # Replace spike bins with background amplitude, keep phase
        for bi in range(max(0, h_idx - cluster_bins),
                        min(len(X), h_idx + cluster_bins + 1)):
            if abs(X[bi]) > threshold:
                X_out[bi] = bg_mean * np.exp(1j * np.angle(X[bi]))

    # ── Step 3: Reconstruct ──────────────────────────────────────────
    ch1_filtered = np.fft.irfft(X_out, n=n)

    # ── Step 4: EMD on residual to handle nonlinear components ───────
    # (simplified — Andrews also applies EMD-BF as a comparison method)
    residual = ch1 - ch1_filtered
    try:
        imfs = _emd_sift(residual, max_imfs=3, max_iter=5)
        # Keep only IMFs whose power is small relative to filtered signal
        # (these are volitional EMG residuals, not artefact)
        emg_power = np.var(ch1_filtered)
        emd_addition = np.zeros(n)
        for imf in imfs:
            if np.var(imf) < 0.5 * emg_power:
                emd_addition += imf
        out = ch1_filtered + emd_addition
    except Exception:
        out = ch1_filtered

    return out, (time.perf_counter() - t0) / n * fs * 1e3


# =============================================================================
#  GUI
# =============================================================================

class App(tk.Tk):
    FS = 4000

    def __init__(self):
        super().__init__()
        self.title("EMG Artefact Suppression Benchmark  v3  —  Ayush Ram · USYD")
        self.configure(bg=BG)
        self.minsize(1220, 740)

        self.loaded_emg       = None
        self.loaded_ch2       = None
        self.loaded_noise_ch1 = None   # rest-segment noise ch1 (dynamic)
        self.loaded_noise_ch2 = None   # rest-segment noise ch2 (dynamic)
        self.loaded_fs        = self.FS
        self.algo_var    = tk.StringVar(value="blank")
        self.tab_var     = tk.StringVar(value="signal")
        self.metric_mode = tk.StringVar(value="inter")
        self.file_label  = tk.StringVar(value="synthetic  ·  no file loaded")
        self.params      = {}
        self._computing  = False
        self.noise_mode  = tk.StringVar(value="synthetic")  # 'synthetic' or 'real'

        self._build()
        self._run()

    # ── main layout ───────────────────────────────────────────────────────────

    def _build(self):
        self.columnconfigure(1, weight=1)
        self.rowconfigure(0, weight=1)

        # ── Scrollable sidebar ────────────────────────────────────────────────
        sb_outer = tk.Frame(self, bg=SURFACE, width=290)
        sb_outer.grid(row=0, column=0, sticky="nsew")
        sb_outer.grid_propagate(False)
        sb_outer.rowconfigure(0, weight=1)
        sb_outer.columnconfigure(0, weight=1)

        sb_canvas = tk.Canvas(sb_outer, bg=SURFACE, highlightthickness=0,
                              width=275)
        sb_canvas.grid(row=0, column=0, sticky="nsew")

        sb_scroll = ttk.Scrollbar(sb_outer, orient="vertical",
                                  command=sb_canvas.yview)
        sb_scroll.grid(row=0, column=1, sticky="ns")
        sb_canvas.configure(yscrollcommand=sb_scroll.set)

        sb = tk.Frame(sb_canvas, bg=SURFACE)
        self._sb_window = sb_canvas.create_window((0, 0), window=sb,
                                                   anchor="nw", width=275)

        def _on_frame_configure(e):
            sb_canvas.configure(scrollregion=sb_canvas.bbox("all"))
        sb.bind("<Configure>", _on_frame_configure)

        # Mouse wheel scrolling
        def _on_mousewheel(e):
            sb_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        sb_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        self._build_sidebar(sb)

        rp = tk.Frame(self, bg=BG)
        rp.grid(row=0, column=1, sticky="nsew")
        rp.rowconfigure(1, weight=1)
        rp.columnconfigure(0, weight=1)

        self._build_metric_strip(rp)

        pf = tk.Frame(rp, bg=BG)
        pf.grid(row=1, column=0, sticky="nsew", padx=10, pady=(6, 4))
        pf.rowconfigure(0, weight=1); pf.columnconfigure(0, weight=1)

        self.fig    = plt.figure(figsize=(10, 6))
        self.canvas = FigureCanvasTkAgg(self.fig, master=pf)
        self.canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")

        self.status_var = tk.StringVar(value="Ready")
        tk.Label(rp, textvariable=self.status_var, font=("Courier", 7),
                 bg=SURFACE, fg=MUTED, anchor="w", padx=10
                 ).grid(row=2, column=0, sticky="ew")

    # ── sidebar ───────────────────────────────────────────────────────────────

    def _build_sidebar(self, sb):
        tk.Label(sb, text="EMG SUPPRESSION", font=("Courier", 11, "bold"),
                 bg=SURFACE, fg=TEXT).pack(anchor="w", padx=18, pady=(18, 0))
        tk.Label(sb, text="Thesis A&B  ·  Ayush Ram  ·  USYD",
                 font=("Courier", 8), bg=SURFACE, fg=MUTED
                 ).pack(anchor="w", padx=18, pady=(1, 0))

        self._div(sb)
        self._sec(sb, "DATA SOURCE")
        tk.Label(sb, textvariable=self.file_label, font=("Courier", 8),
                 bg=SURFACE, fg=ACCENT, wraplength=250, justify="left"
                 ).pack(anchor="w", padx=18, pady=(2, 6))
        bf = tk.Frame(sb, bg=SURFACE); bf.pack(padx=18, fill=tk.X)
        self._btn(bf, "Load file", self._load_file).pack(side=tk.LEFT, padx=(0, 6))
        self._btn(bf, "Clear", self._clear_file).pack(side=tk.LEFT)

        self._div(sb)
        self._sec(sb, "NOISE MODE")
        noise_frame = tk.Frame(sb, bg=SURFACE); noise_frame.pack(fill=tk.X, padx=12, pady=2)
        tk.Radiobutton(
            noise_frame, text="Synthetic  (randn × 5 µV)",
            variable=self.noise_mode, value="synthetic",
            font=("Courier", 8), bg=SURFACE, fg=TEXT,
            selectcolor=CARD, activebackground=SURFACE,
            activeforeground=ACCENT, cursor="hand2",
            command=self._run).pack(anchor="w")
        tk.Label(noise_frame, text="  Always available — static Gaussian",
                 font=("Courier", 7), bg=SURFACE, fg=MUTED).pack(anchor="w")
        tk.Radiobutton(
            noise_frame, text="Real  (Ninapro rest segments)",
            variable=self.noise_mode, value="real",
            font=("Courier", 8), bg=SURFACE, fg=TEXT,
            selectcolor=CARD, activebackground=SURFACE,
            activeforeground=ACCENT, cursor="hand2",
            command=self._run).pack(anchor="w")
        tk.Label(noise_frame, text="  Load a Ninapro .mat file first",
                 font=("Courier", 7), bg=SURFACE, fg=MUTED).pack(anchor="w")

        self._div(sb)
        self._sec(sb, "METRIC MODE")
        self.mode_desc_var = tk.StringVar(value="")
        for label, mode_id, ref in METRIC_MODES:
            f = tk.Frame(sb, bg=SURFACE); f.pack(fill=tk.X, padx=12, pady=1)
            rb = tk.Radiobutton(
                f, text=label, variable=self.metric_mode, value=mode_id,
                font=("Courier", 8, "bold"), bg=SURFACE, fg=TEXT,
                selectcolor=CARD, activebackground=SURFACE,
                activeforeground=ACCENT, cursor="hand2",
                command=self._on_mode_change)
            rb.pack(anchor="w")
            tk.Label(f, text=f"  {ref}", font=("Courier", 7),
                     bg=SURFACE, fg=MUTED).pack(anchor="w")

        self._div(sb)
        self._sec(sb, "NOISE MODE")
        nf = tk.Frame(sb, bg=SURFACE); nf.pack(fill=tk.X, padx=12, pady=2)
        tk.Radiobutton(nf, text="Synthetic  (randn x 5 uV)",
            variable=self.noise_mode, value="synthetic",
            font=("Courier", 8), bg=SURFACE, fg=TEXT, selectcolor=CARD,
            activebackground=SURFACE, activeforeground=ACCENT,
            cursor="hand2", command=self._run).pack(anchor="w")
        tk.Label(nf, text="  Always available - static Gaussian",
            font=("Courier", 7), bg=SURFACE, fg=MUTED).pack(anchor="w")
        tk.Radiobutton(nf, text="Real  (Ninapro rest segments)",
            variable=self.noise_mode, value="real",
            font=("Courier", 8), bg=SURFACE, fg=TEXT, selectcolor=CARD,
            activebackground=SURFACE, activeforeground=ACCENT,
            cursor="hand2", command=self._run).pack(anchor="w")
        tk.Label(nf, text="  Load a Ninapro .mat file first",
            font=("Courier", 7), bg=SURFACE, fg=MUTED).pack(anchor="w")

        self._div(sb)
        self._sec(sb, "PARAMETERS")
        sliders = [
            ("sig_dur_s",  "Synth duration (s)", 1,   10,  2,   1),
            ("stim_freq",  "Stim freq (Hz)",    10,  100, 50,  10),
            ("art_amp_mv", "Artefact (mV)",       5,  300, 50,   5),
            ("emg_amp_uv", "EMG amp (µV RMS)",  100,  800, 300, 100),
            ("mwave_pct",  "M-wave scale (%)",    0,  200, 100,  10),
            ("blank_ms",   "Blank window (ms)",   1,   15,   5,   1),
            ("n_avg",      "Template avg cycles", 3,   30,  10,   1),
            ("ewma_alpha", "EWMA alpha (×0.01)",  1,   50,  10,   1),
            ("lms_mu_x1k", "LMS/eNLMS mu(×.001)",1,   50,  10,   1),
            ("rls_lam_x1k","RLS lam (×0.001)",  990, 999, 999,   1),
        ]
        for args in sliders:
            self._slider(*((sb,) + args))

        self._div(sb)
        self._sec(sb, "ALGORITHM")
        algo_frame = tk.Frame(sb, bg=SURFACE); algo_frame.pack(fill=tk.X)
        for k, name in ALGO_NAMES.items():
            self._algo_radio(algo_frame, k, name)

        self._div(sb)
        self._sec(sb, "VIEW")

        VIEW_TABS = [
            ("signal",       "📈  Signal view",      "Raw · contaminated · output"),
            ("compare",      "🔀  All methods",       "Overlay all 9 algorithms"),
            ("metrics",      "📊  Metrics",           "Bar charts for active mode"),
            ("explainer",    "📖  How it works",      "Selected algo · equations"),
            ("all_explainer","📚  All methods guide", "All 9 — equations + papers"),
            ("metric_info",  "🔬  Metric formula",    "Active mode · paper values"),
            ("algo_metrics", "🧪  All metrics here",  "Every metric for this algo"),
        ]
        for v, lbl, sub in VIEW_TABS:
            f = tk.Frame(sb, bg=SURFACE); f.pack(fill=tk.X, padx=12, pady=1)
            tk.Radiobutton(f, text=lbl, variable=self.tab_var, value=v,
                           font=("Courier", 9, "bold"), bg=SURFACE, fg=TEXT,
                           selectcolor=CARD, activebackground=SURFACE,
                           activeforeground=ACCENT, cursor="hand2",
                           command=self._run).pack(anchor="w")
            tk.Label(f, text=f"  {sub}", font=("Courier", 7),
                     bg=SURFACE, fg=MUTED).pack(anchor="w")

        self._div(sb)
        tk.Button(sb, text="  REGENERATE  ", font=("Courier", 9, "bold"),
                  bg=ACCENT, fg=BG, activebackground="#8aaeff",
                  activeforeground=BG, relief="flat", cursor="hand2",
                  command=self._run).pack(fill=tk.X, padx=18, pady=(0, 6))

        tk.Button(sb, text="  EXPORT TO EXCEL  ", font=("Courier", 9, "bold"),
                  bg="#34d399", fg="#0f1117", activebackground="#6ee7b7",
                  activeforeground="#0f1117", relief="flat", cursor="hand2",
                  command=self._export_excel).pack(fill=tk.X, padx=18, pady=(0, 16))

    # ── metric strip ──────────────────────────────────────────────────────────

    def _build_metric_strip(self, parent):
        self.strip_frame = tk.Frame(parent, bg=SURFACE, height=76)
        self.strip_frame.grid(row=0, column=0, sticky="ew")
        self.strip_frame.pack_propagate(False)
        self.m_vars   = {}
        self.m_labels = {}   # for dynamic relabeling
        self._build_strip_cells("inter")

    def _build_strip_cells(self, mode):
        for w in self.strip_frame.winfo_children():
            w.destroy()
        self.m_vars   = {}
        self.m_labels = {}
        specs = STRIP_LABELS.get(mode, STRIP_LABELS["inter"])
        for i, (key, lbl, w) in enumerate(specs):
            cell = tk.Frame(self.strip_frame, bg=CARD, width=w)
            cell.pack(side=tk.LEFT, fill=tk.Y,
                      padx=(10 if i == 0 else 1, 0), pady=1)
            cell.pack_propagate(False)
            lbl_w = tk.Label(cell, text=lbl, font=("Courier", 7), bg=CARD, fg=MUTED)
            lbl_w.pack(anchor="w", padx=8, pady=(6, 0))
            self.m_labels[key] = lbl_w
            var = tk.StringVar(value="—")
            self.m_vars[key] = var
            tk.Label(cell, textvariable=var, font=("Courier", 12, "bold"),
                     bg=CARD, fg=ACCENT if key == "algo" else TEXT
                     ).pack(anchor="w", padx=8)

    # ── helpers ───────────────────────────────────────────────────────────────

    def _div(self, p):
        tk.Frame(p, bg=BORDER, height=1).pack(fill=tk.X, padx=18, pady=8)

    def _sec(self, p, t):
        tk.Label(p, text=t, font=("Courier", 7, "bold"),
                 bg=SURFACE, fg=MUTED).pack(anchor="w", padx=18, pady=(0, 3))

    def _btn(self, p, text, cmd):
        return tk.Button(p, text=text, font=("Courier", 8),
                         bg=CARD, fg=TEXT, activebackground=BORDER,
                         activeforeground=TEXT, relief="flat",
                         padx=8, pady=4, cursor="hand2", command=cmd)

    def _slider(self, parent, key, label, lo, hi, default, step):
        row = tk.Frame(parent, bg=SURFACE); row.pack(fill=tk.X, padx=18, pady=2)
        tk.Label(row, text=label, font=("Courier", 8), bg=SURFACE,
                 fg=TEXT, width=22, anchor="w").pack(side=tk.LEFT)
        var = tk.IntVar(value=default); self.params[key] = var
        val_lbl = tk.Label(row, text=str(default), font=("Courier", 8, "bold"),
                           bg=SURFACE, fg=ACCENT, width=5, anchor="e")
        val_lbl.pack(side=tk.RIGHT)
        def on(_, v=var, lbl=val_lbl):
            lbl.config(text=str(v.get())); self._run()
        ttk.Scale(row, from_=lo, to=hi, variable=var, orient=tk.HORIZONTAL,
                  command=on).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=6)

    def _algo_radio(self, parent, value, text):
        f = tk.Frame(parent, bg=SURFACE); f.pack(fill=tk.X, padx=18, pady=1)
        c = tk.Canvas(f, width=8, height=8, bg=SURFACE, highlightthickness=0)
        col = ALGO_COLORS[value]
        c.create_oval(1, 1, 7, 7, fill=col, outline=col)
        c.pack(side=tk.LEFT, padx=(0, 6))
        tk.Radiobutton(f, text=text, variable=self.algo_var, value=value,
                       font=("Courier", 9), bg=SURFACE, fg=TEXT,
                       selectcolor=SURFACE, activebackground=SURFACE,
                       activeforeground=TEXT, cursor="hand2",
                       command=self._run).pack(side=tk.LEFT)

    def _on_mode_change(self):
        mode = self.metric_mode.get()
        self._build_strip_cells(mode)
        self._run()

    # ── file I/O ──────────────────────────────────────────────────────────────

    def _load_file(self):
        path = filedialog.askopenfilename(
            filetypes=[("EMG files", "*.mat *.csv *.npy *.txt"), ("All", "*.*")])
        if not path: return
        try:
            data, actual_fs, info, ch2_data, noise_ch1, noise_ch2 =                 load_emg_file(path, self.FS)
            self.loaded_emg       = data
            self.loaded_ch2       = ch2_data
            self.loaded_noise_ch1 = noise_ch1
            self.loaded_noise_ch2 = noise_ch2
            self.loaded_fs        = actual_fs
            ch2_note   = ("  [real ch2]" if ch2_data is not None
                          else "  [synth ch2]")
            noise_note = ("  [dynamic noise]" if noise_ch1 is not None
                          else "  [static noise]")
            self.file_label.set(info + ch2_note + noise_note)
            self._run()
        except RuntimeError as e:
            messagebox.showerror("Load error", str(e))

    def _clear_file(self):
        self.loaded_emg       = None
        self.loaded_ch2       = None
        self.loaded_noise_ch1 = None
        self.loaded_noise_ch2 = None
        self.loaded_fs        = self.FS
        self.file_label.set("synthetic  ·  no file loaded")
        self._run()

    def _get_noise_segment(self, n, channel=1):
        """
        Return n samples of background noise for ch1 or ch2.

        Synthetic mode: randn * 5e-6  (static 5 uV Gaussian — original)

        Real mode: rest segments from the loaded Ninapro file, tiled with
        a random phase offset. Scaled to REAL_NOISE_UV (default 20 uV RMS)
        — realistic for surface EMG baseline (skin-electrode impedance,
        amplifier noise, breathing artifacts). This is visible on the plot
        and meaningfully different from the flat Gaussian floor.
        No rescaling to 5 uV — that was defeating the purpose.
        """
        REAL_NOISE_UV = 20e-6   # 20 uV RMS — realistic sEMG noise floor

        if self.noise_mode.get() == "real":
            src = (self.loaded_noise_ch1 if channel == 1
                   else self.loaded_noise_ch2)
            if src is not None and len(src) > 0:
                reps   = int(np.ceil(n / len(src)))
                tiled  = np.tile(src, reps)[:n]
                offset = np.random.randint(0, len(src))
                tiled  = np.roll(tiled, offset)
                # src is stored at unit std — scale to target noise amplitude
                tiled  = tiled * REAL_NOISE_UV
                return tiled
            # Real selected but no file loaded — fall back silently
            return np.random.randn(n) * 5e-6
        else:
            # Synthetic mode — static Gaussian, unchanged from original
            return np.random.randn(n) * 5e-6

    # ── compute ───────────────────────────────────────────────────────────────

    def _get_params(self):
        return dict(
            sig_dur_s  = self.params["sig_dur_s"].get(),        # synthetic only
            stim_freq  = self.params["stim_freq"].get(),
            art_amp    = self.params["art_amp_mv"].get() * 1e-3,
            emg_amp    = self.params["emg_amp_uv"].get() * 1e-6,
            mwave_pct  = self.params["mwave_pct"].get() / 100.0,
            blank_ms   = self.params["blank_ms"].get(),
            n_avg      = self.params["n_avg"].get(),
            ewma_alpha = self.params["ewma_alpha"].get() / 100,
            lms_mu     = self.params["lms_mu_x1k"].get() / 1000,
            rls_lam    = self.params["rls_lam_x1k"].get() / 1000,
        )

    def _compute(self, p, full=False):
        fs    = self.loaded_fs if self.loaded_emg is not None else self.FS
        if self.loaded_emg is not None:
            # Real file loaded — use file length, ignore duration slider
            clean = self.loaded_emg.copy()
            n = len(clean)
            max_samples = int(fs * 30.0)   # allow up to 30s for real data
            if n > max_samples: clean = clean[:max_samples]; n = max_samples
        else:
            # Synthetic — use the duration slider (default 2s, up to 10s)
            dur = float(p.get("sig_dur_s", 2))
            clean = gen_emg(fs, dur, p["emg_amp"])
            n = len(clean)

        art, times, interval = gen_artefact(
            fs, n, p["stim_freq"], p["art_amp"],
            emg_amp=p["emg_amp"] * p["mwave_pct"])
        ch1 = clean + art + self._get_noise_segment(n, channel=1)

        # Use real second channel if available from loaded file,
        # otherwise fall back to synthetic make_ch2().
        # Real ch2 has genuine muscle signal morphology with ~3% EMG
        # cross-correlation (measured from NinaPro DB3 adjacent electrodes),
        # which is more faithful than synthetic Gaussian noise.
        if self.loaded_ch2 is not None:
            raw_ch2 = self.loaded_ch2[:n].copy()
            # Add same synthetic artefact (FES field spreads uniformly
            # across adjacent electrodes — 92% amplitude correlation)
            art_ch2 = art * (0.92 + np.random.randn() * 0.02)
            ch2 = raw_ch2 + art_ch2 + self._get_noise_segment(n, channel=2)
        else:
            ch2 = make_ch2(clean, art, fs, emg_amp=p["emg_amp"] * p["mwave_pct"])

        bms = p["blank_ms"]; mu = p["lms_mu"]; lam = p["rls_lam"]
        na  = p["n_avg"];    al = p["ewma_alpha"]
        sf  = p["stim_freq"]; algo = self.algo_var.get()
        mode = self.metric_mode.get()

        def run_one(k):
            dispatch = {
                "blank"   : lambda: algo_blanking(ch1, times, fs, bms),
                "template": lambda: algo_fixed_template(ch1, times, interval, fs, na),
                "ewma"    : lambda: algo_ewma_template(ch1, times, interval, fs, al),
                "destd"   : lambda: algo_destd(ch1, ch2, times, interval, fs),
                "gso"     : lambda: algo_gso(ch1, ch2, times, interval, fs),
                "lms"     : lambda: algo_lms(ch1, ch2, times, interval, fs, mu),
                "enlms"   : lambda: algo_enlms(ch1, ch2, times, interval, fs, mu),
                "rls"     : lambda: algo_rls(ch1, ch2, times, interval, fs, lam),
                "ceemdan" : lambda: algo_ceemdan(ch1, times, interval, fs, sf),
            }
            return dispatch[k]()

        def metric_one(k, out_sig):
            return compute_metrics(
                clean, out_sig, k, bms, sf,
                times=times, fs=fs, mode=mode, art=art)

        if full:
            results = {k: run_one(k) for k in ALGO_NAMES}
            out, lat = results[algo]
            all_out  = {k: v[0] for k, v in results.items()}
            all_lats = {k: v[1] for k, v in results.items()}
            all_m    = {k: metric_one(k, all_out[k]) for k in ALGO_NAMES}
            for k in all_m: all_m[k]["latency"] = all_lats[k]
        else:
            out, lat = run_one(algo)
            all_out  = {algo: out}
            all_lats = {algo: lat}
            all_m    = {algo: metric_one(algo, out)}
            all_m[algo]["latency"] = lat

        m = metric_one(algo, out)
        return dict(clean=clean, art=art, ch1=ch1, ch2=ch2,
                    out=out, lat=lat, algo=algo, times=times,
                    all_out=all_out, m=m, all_m=all_m, fs=fs, n=n,
                    mode=mode)

    # ── run ───────────────────────────────────────────────────────────────────

    def _run(self, *_):
        if hasattr(self, "_run_after_id"):
            self.after_cancel(self._run_after_id)
        self._run_after_id = self.after(280, self._run_now)

    def _run_now(self):
        if self._computing: return
        self._computing = True
        tab = self.tab_var.get()
        need_full = tab in ("compare", "metrics", "algo_metrics")
        self.status_var.set("Computing all 9 algorithms…" if need_full else "Computing…")
        self.update_idletasks()
        try:
            p = self._get_params()
            d = self._compute(p, full=need_full)
        except Exception as e:
            messagebox.showerror("Error", str(e)); self._computing = False; return

        m    = d["m"]; algo = d["algo"]; mode = d["mode"]
        fmts = STRIP_FORMATS.get(mode, STRIP_FORMATS["inter"])

        self.m_vars["algo"].set(ALGO_NAMES[algo])

        def _fmt(key, val):
            fmt = fmts.get(key, "—")
            if fmt == "—": return "—"
            try: return fmt.format(val)
            except: return "—"

        self.m_vars["snr"].set(_fmt("snr", m["snr"]))
        self.m_vars["sdr"].set(_fmt("sdr", m["sdr"]))
        self.m_vars["rmse"].set(_fmt("rmse", m["rmse"]))
        self.m_vars["r"].set(_fmt("r", m["r"]))
        self.m_vars["loss"].set(
            _fmt("loss", m["loss"]) if m["loss"] else "0%")
        self.m_vars["lat"].set(f"{d['lat']:.4f} ms")

        self._draw(d)
        fs_disp = self.loaded_fs if self.loaded_emg is not None else self.FS
        mode_label = next((lbl for lbl, mid, _ in METRIC_MODES if mid == mode), mode)
        dur_str = (f"{d['n']/fs_disp:.1f}s [file]"
                   if self.loaded_emg is not None
                   else f"{d['n']/fs_disp:.1f}s [synth]")
        self.status_var.set(
            f"fs={fs_disp} Hz  dur={dur_str}  stim={p['stim_freq']} Hz  "
            f"art={p['art_amp']*1000:.0f} mV  EMG={p['emg_amp']*1e6:.0f} µV  "
            f"algo={ALGO_NAMES[algo]}  lat={d['lat']:.4f} ms  "
            f"metric={mode_label}"
        )
        self._computing = False

    # ── drawing ───────────────────────────────────────────────────────────────

    @staticmethod
    def _zoom(arr, fs, secs=0.35):
        nz = min(int(fs * secs), len(arr))
        return np.arange(nz) / fs * 1000, arr[:nz] * 1e6

    @staticmethod
    def _vlines(ax, times, fs, nz):
        lim = nz / fs * 1000
        for st in times:
            if st / fs * 1000 <= lim:
                ax.axvline(st / fs * 1000, color="#ef4444",
                           lw=0.5, alpha=0.2, zorder=0)

    @staticmethod
    def _style(ax, title, ylabel="µV", xlabel=None):
        ax.set_title(title, pad=4, fontsize=8.5)
        ax.set_ylabel(ylabel, labelpad=3)
        if xlabel: ax.set_xlabel(xlabel, labelpad=3)
        ax.spines[["top", "right"]].set_visible(False)
        ax.spines[["left", "bottom"]].set_color(BORDER)

    def _draw(self, d):
        self.fig.clf()
        tab = self.tab_var.get()
        {"signal"       : self._draw_signal,
         "compare"      : self._draw_compare,
         "metrics"      : self._draw_metrics,
         "explainer"    : self._draw_explainer,
         "all_explainer": self._draw_all_explainer,
         "metric_info"  : self._draw_metric_info,
         "algo_metrics" : self._draw_algo_metrics}[tab](d)
        self.fig.tight_layout(pad=2.2)
        self.canvas.draw()

    def _draw_signal(self, d):
        fs = d["fs"]; nz = min(int(fs * 0.35), d["n"])
        gs = GridSpec(3, 1, figure=self.fig, hspace=0.55)
        rows = [
            (d["clean"], "#34d399", "Ground truth EMG  (reference)"),
            (d["ch1"],   MUTED,     "Contaminated channel 1"),
            (d["out"],   ALGO_COLORS[d["algo"]],
             f"{ALGO_NAMES[d['algo']]}  output"),
        ]
        for i, (sig, col, title) in enumerate(rows):
            ax = self.fig.add_subplot(gs[i])
            t, s = self._zoom(sig, fs)
            ax.plot(t, s, color=col, lw=0.9, label=title)
            if i == 2:
                tr, cr = self._zoom(d["clean"], fs)
                ax.plot(tr, cr, color="#34d399", lw=0.6,
                        alpha=0.35, ls="--", label="reference")
                ax.legend(loc="upper right")
            self._vlines(ax, d["times"], fs, nz)
            self._style(ax, title, xlabel="Time (ms)" if i == 2 else None)

    def _draw_compare(self, d):
        fs = d["fs"]; nz = min(int(fs * 0.35), d["n"])
        ax = self.fig.add_subplot(111)
        t, c = self._zoom(d["clean"], fs)
        ax.plot(t, c, color="#34d399", lw=1.3, ls="--",
                alpha=0.5, label="Clean EMG", zorder=9)
        for k, out in d["all_out"].items():
            t2, s = self._zoom(out, fs)
            ax.plot(t2, s, color=ALGO_COLORS[k], lw=0.75,
                    label=ALGO_NAMES[k], alpha=0.85)
        self._vlines(ax, d["times"], fs, nz)
        self._style(ax, "All 9 algorithms — first 350 ms", xlabel="Time (ms)")
        ax.legend(loc="upper right", ncol=3)

    def _draw_metrics(self, d):
        am    = d["all_m"]
        algos = list(am.keys())
        cols  = [ALGO_COLORS[k] for k in algos]
        mode  = d["mode"]

        # Choose which values to plot depending on metric mode
        if mode == "inter":
            panels = [
                ("SNR (dB)",       [am[k]["snr"]     for k in algos], "dB"),
                ("SDR (dB)",       [am[k]["sdr"]     for k in algos], "dB"),
                ("RMSE (µV)",      [am[k]["rmse"]    for k in algos], "µV"),
                ("Correlation r",  [am[k]["r"]       for k in algos], ""),
                ("Data loss (%)",  [am[k]["loss"]    for k in algos], "%"),
                ("Latency (ms/s)", [am[k]["latency"] for k in algos], "ms"),
            ]
        elif mode == "wang":
            panels = [
                ("SNR dB (Wang)",  [am[k]["snr"]     for k in algos], "dB"),
                ("NRMSE",          [am[k]["nrmse"]   for k in algos], ""),
                ("RMSE (µV)",      [am[k]["rmse"]    for k in algos], "µV"),
                ("Correlation r",  [am[k]["r"]       for k in algos], ""),
                ("Data loss (%)",  [am[k]["loss"]    for k in algos], "%"),
                ("Latency (ms/s)", [am[k]["latency"] for k in algos], "ms"),
            ]
        elif mode == "sennels":
            panels = [
                ("MRI_y (dB)",     [am[k]["mri_y"]   for k in algos], "dB"),
                ("PR (dB)",        [am[k]["pr"]       for k in algos], "dB"),
                ("RMSE (µV)",      [am[k]["rmse"]    for k in algos], "µV"),
                ("Correlation r",  [am[k]["r"]       for k in algos], ""),
                ("Data loss (%)",  [am[k]["loss"]    for k in algos], "%"),
                ("Latency (ms/s)", [am[k]["latency"] for k in algos], "ms"),
            ]
        elif mode == "asr":
            panels = [
                ("ASR (dB)",       [am[k]["asr"]     for k in algos], "dB"),
                ("RMSE (µV)",      [am[k]["rmse"]    for k in algos], "µV"),
                ("Correlation r",  [am[k]["r"]       for k in algos], ""),
                ("Data loss (%)",  [am[k]["loss"]    for k in algos], "%"),
                ("Latency (ms/s)", [am[k]["latency"] for k in algos], "ms"),
                ("SNR (dB) ref",   [am[k]["snr"]     for k in algos], "dB"),
            ]
        elif mode == "mandrile":
            panels = [
                ("ARV_norm (%)",   [am[k]["arv_norm"] for k in algos], "%"),
                ("RMSE (µV)",      [am[k]["rmse"]     for k in algos], "µV"),
                ("Correlation r",  [am[k]["r"]        for k in algos], ""),
                ("Data loss (%)",  [am[k]["loss"]      for k in algos], "%"),
                ("Latency (ms/s)", [am[k]["latency"]  for k in algos], "ms"),
                ("SNR (dB) ref",   [am[k]["snr"]      for k in algos], "dB"),
            ]
        elif mode == "limnuson":
            panels = [
                ("RMS reduction ×", [am[k]["rms_red"]  for k in algos], "×"),
                ("M-wave CC",       [am[k]["sdr"]      for k in algos], ""),
                ("M-wave RMSE µV",  [am[k]["rmse"]     for k in algos], "µV"),
                ("Correlation r",   [am[k]["r"]        for k in algos], ""),
                ("Data loss (%)",   [am[k]["loss"]     for k in algos], "%"),
                ("Latency (ms/s)",  [am[k]["latency"]  for k in algos], "ms"),
            ]
        elif mode == "andrews":
            panels = [
                ("FFT% vs input",   [am[k]["fft_pct"]  for k in algos], "%"),
                ("FFT% vs clean",   [am[k]["sdr"]      for k in algos], "%"),
                ("EMGrms ratio",    [am[k]["nrmse"]    for k in algos], "×"),
                ("RMSE (µV)",       [am[k]["rmse"]     for k in algos], "µV"),
                ("Correlation r",   [am[k]["r"]        for k in algos], ""),
                ("Latency (ms/s)",  [am[k]["latency"]  for k in algos], "ms"),
            ]
        else:
            panels = []

        gs = GridSpec(2, 3, figure=self.fig, hspace=0.65, wspace=0.42)
        for idx, (title, vals, unit) in enumerate(panels[:6]):
            ax = self.fig.add_subplot(gs[idx // 3, idx % 3])
            bars = ax.bar(range(len(algos)), vals, color=cols,
                          edgecolor=BORDER, linewidth=0.5, width=0.6)
            self._style(ax, title, ylabel=unit)
            ax.set_xticks(range(len(algos)))
            ax.set_xticklabels([ALGO_NAMES[k][:6] for k in algos],
                               fontsize=6, rotation=35, ha="right")
            vmax = max(abs(v) for v in vals) if vals else 1
            for bar, v in zip(bars, vals):
                ax.text(bar.get_x() + bar.get_width() / 2,
                        bar.get_height() + vmax * 0.015,
                        f"{v:.1f}", ha="center", va="bottom",
                        fontsize=6, color=MUTED)

        # Mode watermark
        mode_label = next((lbl for lbl, mid, _ in METRIC_MODES if mid == mode), mode)
        self.fig.text(0.98, 0.02, f"Metric: {mode_label}",
                      ha="right", va="bottom", fontsize=7,
                      color=ACCENT, alpha=0.6, fontfamily="monospace")

    def _draw_explainer(self, d):
        algo = d["algo"]; ex = EXPLAINERS[algo]
        ax = self.fig.add_subplot(111)
        ax.set_facecolor(BG); ax.axis("off")
        y = 0.97
        ax.text(0.04, y, ex["title"], transform=ax.transAxes,
                fontsize=15, fontweight="bold", va="top", color=TEXT,
                fontfamily="monospace")
        y -= 0.08
        ax.text(0.04, y, ex["ref"], transform=ax.transAxes,
                fontsize=8.5, va="top", color=MUTED, fontfamily="monospace")
        y -= 0.10
        ax.text(0.04, y, f"  {ex['eq']}  ", transform=ax.transAxes,
                fontsize=11, va="top", color=ACCENT, fontfamily="monospace",
                bbox=dict(boxstyle="round,pad=0.5", facecolor=CARD,
                          edgecolor=ACCENT, linewidth=1.2))
        y -= 0.14
        ax.text(0.04, y, "\n".join(ex["lines"]),
                transform=ax.transAxes, fontsize=9.5, va="top",
                color=TEXT, fontfamily="monospace", linespacing=1.85)
        ax.text(0.96, 0.04, ALGO_NAMES[algo].upper(),
                transform=ax.transAxes, fontsize=24, fontweight="bold",
                va="bottom", ha="right", color=ALGO_COLORS[algo],
                alpha=0.10, fontfamily="monospace")

    def _draw_all_explainer(self, d):
        """
        All-methods guide: 3×3 grid, one panel per algorithm.
        Each panel shows: name · equation · 2-line summary · paper ref · colour bar.
        """
        keys   = list(ALGO_NAMES.keys())   # 9 algos
        gs     = GridSpec(3, 3, figure=self.fig,
                          hspace=0.18, wspace=0.12,
                          left=0.02, right=0.99,
                          top=0.91, bottom=0.02)

        self.fig.text(0.5, 0.975, "All 9 Algorithms — How They Work",
                      ha="center", va="top", fontsize=11, fontweight="bold",
                      color=TEXT, fontfamily="monospace")

        COMPACT = {
            "blank"   : ("D = Tb × fs",
                         "Zero signal for fixed\nwindow after each pulse.\nSimplest hardware gate.",
                         "Mandrile 2003 · Huang 2023"),
            "template": ("y = x − s_avg",
                         "Subtract cycle-averaged\nartefact template.\nCalibration then fixed.",
                         "Limnuson 2014 · Liu 2025"),
            "ewma"    : ("s_{k+1}=(1-α)s_k + α·x_k",
                         "IIR recursive template,\ntracks slow drift.\nWarm-start from pulse 1.",
                         "Limnuson 2014 · §2.6.2 Eq.6"),
            "destd"   : ("v = (Δx1 − Δx2) / 2",
                         "Spatial subtraction:\nartefact shared across\nchannels, EMG differs.",
                         "Mandrile 2003 · Chen 2023"),
            "gso"     : ("z = x − Σ<x,qk>qk",
                         "Project onto reference\nbasis, subtract artefact\nsubspace (6 taps).",
                         "Chen 2023 · §2.5.2 Eq.5"),
            "lms"     : ("w[n+1] = w[n] + μ·e·x",
                         "Gradient descent on\nsquared error per sample.\nO(N) compute.",
                         "Sennels 1997 · §2.7.1 Eq.7"),
            "enlms"   : ("w += μ/(‖x‖²+ε)·e·x",
                         "Normalised LMS: step\nscales with input power.\nStable for non-stationary.",
                         "Boyer 2023 · §2.7.1 Eq.8"),
            "rls"     : ("k = Px/(λ + x'Px)",
                         "Minimise total past\nsquared error. O(N²)\nbut fast convergence.",
                         "Sennels 1997 · §2.7.1 Eq.9-11"),
            "ceemdan" : ("x = Σ IMFk  [art filtered]",
                         "Decompose into IMFs,\nidentify artefact ones,\nfilter & reconstruct.",
                         "Andrews 2023 · §2.7.2"),
        }

        import matplotlib.patches as mpatches

        for idx, key in enumerate(keys):
            row, col = divmod(idx, 3)
            ax = self.fig.add_subplot(gs[row, col])
            ax.set_facecolor(CARD)
            ax.set_xlim(0, 1); ax.set_ylim(0, 1)
            ax.axis("off")

            color = ALGO_COLORS[key]
            eq, summary, ref = COMPACT[key]

            # Colour accent bar — left edge
            bar = mpatches.Rectangle(
                (0, 0), 0.022, 1.0,
                facecolor=color, edgecolor="none",
                transform=ax.transAxes, zorder=5, clip_on=False)
            ax.add_patch(bar)

            # Algorithm name
            ax.text(0.06, 0.93, ALGO_NAMES[key],
                    transform=ax.transAxes, fontsize=8.5, fontweight="bold",
                    va="top", color=color, fontfamily="monospace")

            # Number watermark
            ax.text(0.97, 0.95, f"#{idx+1}",
                    transform=ax.transAxes, fontsize=18, fontweight="bold",
                    va="top", ha="right", color=color, alpha=0.13,
                    fontfamily="monospace")

            # Divider line under name
            ax.axhline(0.80, xmin=0.05, xmax=0.95,
                       color=BORDER, linewidth=0.6)

            # Equation — highlighted box
            ax.text(0.06, 0.77, eq,
                    transform=ax.transAxes, fontsize=7.5, va="top",
                    color=ACCENT, fontfamily="monospace",
                    bbox=dict(boxstyle="round,pad=0.25", facecolor=BG,
                              edgecolor=color, linewidth=0.8, alpha=0.9))

            # Summary — 3 lines
            ax.text(0.06, 0.54, summary,
                    transform=ax.transAxes, fontsize=7.5, va="top",
                    color=TEXT, fontfamily="monospace", linespacing=1.55)

            # Divider above ref
            ax.axhline(0.12, xmin=0.05, xmax=0.95,
                       color=BORDER, linewidth=0.5)

            # Paper reference — bottom
            ax.text(0.06, 0.09, ref,
                    transform=ax.transAxes, fontsize=6.5, va="top",
                    color=MUTED, fontfamily="monospace")

    def _draw_algo_metrics(self, d):
        """
        Show ALL 7 paper metrics for the currently selected algorithm.
        Computes each metric mode in turn and displays as a dashboard.
        Each row = one paper, columns = metric values.
        """
        algo  = d["algo"]
        clean = d["clean"]
        art   = d["art"]
        rec   = d["out"]
        times = d["times"]
        fs    = d["fs"]
        p     = self._get_params()
        bms   = p["blank_ms"]
        sf    = p["stim_freq"]

        # Compute all 7 metric modes for this one algorithm
        results = {}
        for _, mode_id, _ in METRIC_MODES:
            try:
                results[mode_id] = compute_metrics(
                    clean, rec, algo, bms, sf,
                    times=times, fs=fs, mode=mode_id, art=art)
            except Exception:
                results[mode_id] = {}

        ax = self.fig.add_subplot(111)
        ax.set_facecolor(BG); ax.axis("off")

        color = ALGO_COLORS[algo]
        ax.text(0.02, 0.97, f"All Paper Metrics — {ALGO_NAMES[algo]}",
                transform=ax.transAxes, fontsize=13, fontweight="bold",
                va="top", color=color, fontfamily="monospace")
        ax.text(0.02, 0.91, "Latency: " + f"{d['lat']:.4f} ms/sample",
                transform=ax.transAxes, fontsize=9, va="top",
                color=MUTED, fontfamily="monospace")

        # Table definition: (mode_id, paper, metric_name, field, unit, target)
        rows = [
            ("inter",    "Your §2.10",      "SNR",         "snr",     "dB",  "higher = better"),
            ("inter",    "Your §2.10",      "RMSE",        "rmse",    "µV",  "lower = better"),
            ("inter",    "Your §2.10",      "Pearson r",   "r",       "",    "→ 1.0"),
            ("wang",     "Wang/Chen 21/23", "SNR (full)",  "snr",     "dB",  "−15 to −46 dB"),
            ("wang",     "Wang/Chen 21/23", "NRMSE",       "nrmse",   "",    "2.28–11.35"),
            ("sennels",  "Sennels 1997",    "MRI_y",       "mri_y",   "dB",  "→ 0 dB"),
            ("sennels",  "Sennels 1997",    "PR",          "pr",      "dB",  "20–28 dB"),
            ("asr",      "Liu 2025",        "ASR",         "asr",     "dB",  "> 20 dB"),
            ("mandrile", "Mandrile 2003",   "ARV_norm",    "arv_norm","%",   "< 46%"),
            ("limnuson", "Limnuson 2014",   "RMS reduc.",  "rms_red", "×",   "> 5×"),
            ("limnuson", "Limnuson 2014",   "Mwave CC",    "sdr",     "",    "→ 0.89"),
            ("limnuson", "Limnuson 2014",   "Mwave RMSE",  "rmse",    "µV",  "< 223 µV"),
            ("andrews",  "Andrews 2023",    "FFT% (input)","fft_pct", "%",   "~96%"),
        ]

        # Header
        y = 0.84
        col_x = [0.02, 0.22, 0.38, 0.55, 0.68, 0.82]
        headers = ["Paper", "Metric", "Value", "Unit", "Target"]
        for hdr, x in zip(headers, col_x[1:]):
            ax.text(x, y, hdr, transform=ax.transAxes,
                    fontsize=8, fontweight="bold", va="top",
                    color=ACCENT2, fontfamily="monospace")
        import matplotlib.patches as mp
        ax.add_patch(mp.FancyArrowPatch(
            (0.01, y - 0.03), (0.99, y - 0.03),
            transform=ax.transAxes, color=BORDER,
            linewidth=0.8, arrowstyle="-"))
        y -= 0.055

        prev_paper = ""
        for mode_id, paper, metric_name, field, unit, target in rows:
            m = results.get(mode_id, {})
            val = m.get(field, None)

            # colour by paper group
            paper_colors = {
                "Your §2.10"     : "#94a3b8",
                "Wang/Chen 21/23": "#38bdf8",
                "Sennels 1997"   : "#a78bfa",
                "Liu 2025"       : "#34d399",
                "Mandrile 2003"  : "#6b7280",
                "Limnuson 2014"  : "#f97316",
                "Andrews 2023"   : "#fb7185",
            }
            pc = paper_colors.get(paper, MUTED)

            # paper label only on first row of each paper
            paper_label = paper if paper != prev_paper else ""
            prev_paper  = paper

            ax.text(col_x[0], y, paper_label, transform=ax.transAxes,
                    fontsize=7.5, va="top", color=pc, fontfamily="monospace",
                    fontweight="bold")
            ax.text(col_x[1], y, metric_name, transform=ax.transAxes,
                    fontsize=8, va="top", color=TEXT, fontfamily="monospace")

            if val is not None:
                try:
                    val_str = f"{val:.2f}"
                except Exception:
                    val_str = str(val)
            else:
                val_str = "—"

            ax.text(col_x[2], y, val_str, transform=ax.transAxes,
                    fontsize=8, va="top", color=color,
                    fontfamily="monospace", fontweight="bold")
            ax.text(col_x[3], y, unit, transform=ax.transAxes,
                    fontsize=7.5, va="top", color=MUTED, fontfamily="monospace")
            ax.text(col_x[4], y, target, transform=ax.transAxes,
                    fontsize=7, va="top", color=MUTED, fontfamily="monospace")

            y -= 0.052
            if y < 0.03:
                break

        # Watermark
        ax.text(0.97, 0.04, ALGO_NAMES[algo].upper(),
                transform=ax.transAxes, fontsize=28, fontweight="bold",
                va="bottom", ha="right", color=color, alpha=0.08,
                fontfamily="monospace")

    def _draw_metric_info(self, d):
        """Dedicated view showing the active metric mode's formula and rationale."""
        mode = d["mode"]
        desc = METRIC_DESCRIPTIONS.get(mode, "No description available.")
        mode_label = next((lbl for lbl, mid, _ in METRIC_MODES if mid == mode), mode)
        ax = self.fig.add_subplot(111)
        ax.set_facecolor(BG); ax.axis("off")

        # Title
        ax.text(0.04, 0.96, "METRIC MODE", transform=ax.transAxes,
                fontsize=9, fontweight="bold", va="top", color=MUTED,
                fontfamily="monospace")
        ax.text(0.04, 0.90, mode_label, transform=ax.transAxes,
                fontsize=17, fontweight="bold", va="top", color=ACCENT,
                fontfamily="monospace")

        # Formula box
        ax.text(0.04, 0.78, desc, transform=ax.transAxes,
                fontsize=10, va="top", color=TEXT, fontfamily="monospace",
                linespacing=2.0,
                bbox=dict(boxstyle="round,pad=0.7", facecolor=CARD,
                          edgecolor=ACCENT, linewidth=1.0, alpha=0.8))

        # Comparison table header
        y_tbl = 0.35
        ax.text(0.04, y_tbl, "PAPER-REPORTED REFERENCE VALUES",
                transform=ax.transAxes, fontsize=8, fontweight="bold",
                va="top", color=MUTED, fontfamily="monospace")
        y_tbl -= 0.06

        ref_table = {
            "inter": [
                ("All methods",  "Your framework", "§2.10", "Unified benchmark"),
            ],
            "wang": [
                ("GS-APEF",  "SNR −15 to −46 dB", "Wang 2021", "r=4–20 FES/EMG"),
                ("LMS-AF",   "SNR similar to GS",  "Wang 2021", "No sig. difference"),
                ("GSO",      "NRMSE 2.28–11.35",   "Chen 2023", "6th order"),
            ],
            "sennels": [
                ("Fixed (comb)", "PR ~14–19 dB",  "Sennels 1997", "Baseline"),
                ("LMS M=1",      "PR ~14–20 dB",  "Sennels 1997", "Adaptive"),
                ("LMS M=6",      "PR ~20–28 dB",  "Sennels 1997", "Best adaptive"),
                ("MRI_y target", "≈ 0 dB",        "Sennels 1997", "= volitional power"),
            ],
            "asr": [
                ("Blanking",     "ASR = ∞ (theoretical)", "Liu 2025", "Data loss trade-off"),
                ("Pole-shifting","ASR ≈ 6 dB",    "Liu 2025", "Simple HW"),
                ("Impedance tmpl","ASR > 20 dB",  "Liu 2025", "Randles model FPGA"),
                ("IIR EWMA",     "IIR converges", "Limnuson 2014", "FPGA impl."),
            ],
            "mandrile": [
                ("Far electrode",  "ARV_norm ~46%",  "Mandrile 2003", "Min distance"),
                ("Near electrode", "ARV_norm ~91%",  "Mandrile 2003", "Max distance"),
                ("Double diff",    "Similar to SD",  "Mandrile 2003", "Spatial filter"),
                ("Waveform shape", "No sig. effect", "Mandrile 2003", "Hybrid stimulator"),
            ],
            "limnuson": [
                ("IIR EWMA",      "17× RMS reduc.",  "Limnuson 2014", "Aplysia dataset"),
                ("IIR EWMA",      "5.3× RMS reduc.", "Limnuson 2014", "Rat cortex"),
                ("IIR EWMA",      "<0.5 ms latency", "Limnuson 2014", "Neural recovery"),
                ("Liu 2014 tmpl", "CC 0.64→0.89",    "Liu 2014 nihms","M-wave CC"),
                ("Liu 2014 tmpl", "425→223 µV RMSE", "Liu 2014 nihms","M-wave RMSE"),
            ],
            "andrews": [
                ("AA-IF",         "96±5% FFT kept",  "Andrews 2023",  "vs EMD-BF"),
                ("EMD-BF",        "75±6% FFT kept",  "Andrews 2023",  "Baseline"),
                ("AA-IF",         "p<0.001 better",  "Andrews 2023",  "Statistical"),
                ("CEEMDAN-IT",    "Best SNR @ low",  "Boyer 2023",    "vs EMD, SWT"),
            ],
        }

        rows = ref_table.get(mode, [])
        col_x = [0.04, 0.28, 0.52, 0.72]
        headers = ["Algorithm", "Value", "Paper", "Note"]
        for j, (hdr, x) in enumerate(zip(headers, col_x)):
            ax.text(x, y_tbl, hdr, transform=ax.transAxes,
                    fontsize=7, fontweight="bold", va="top",
                    color=ACCENT2, fontfamily="monospace")
        y_tbl -= 0.05

        for row in rows:
            for j, (cell, x) in enumerate(zip(row, col_x)):
                ax.text(x, y_tbl, cell, transform=ax.transAxes,
                        fontsize=7.5, va="top", color=TEXT,
                        fontfamily="monospace")
            y_tbl -= 0.045

        # Watermark
        ax.text(0.96, 0.04, mode.upper(),
                transform=ax.transAxes, fontsize=28, fontweight="bold",
                va="bottom", ha="right", color=ACCENT, alpha=0.08,
                fontfamily="monospace")


    # ── Excel export ──────────────────────────────────────────────────────────

    def _export_excel(self):
        """Called when the Export button is clicked."""
        import tkinter.filedialog as fd
        import datetime
        path = fd.asksaveasfilename(
            title="Save benchmark results",
            defaultextension=".xlsx",
            initialfile=f"emg_benchmark_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")]
        )
        if not path:
            return
        self.status_var.set("Running all 9 × 7 metric combinations — please wait…")
        self.update_idletasks()
        try:
            self._do_export(path)
            self.status_var.set(f"Exported → {os.path.basename(path)}")
            messagebox.showinfo("Export complete", f"Saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            self.status_var.set("Export failed — see error dialog")

    def _do_export(self, path):
        """
        Runs every algorithm × every metric mode on current data,
        then writes a formatted 4-sheet .xlsx workbook:
          Sheet 1 — All Metrics   (master table, best values highlighted)
          Sheet 2 — Raw Data      (one row per algo × mode)
          Sheet 3 — Literature    (paper reference values)
          Sheet 4 — Params        (parameter snapshot for this run)
        """
        import datetime
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl not installed. Run:  pip install openpyxl")

        p  = self._get_params()
        fs = self.loaded_fs if self.loaded_emg is not None else self.FS
        if self.loaded_emg is not None:
            clean = self.loaded_emg.copy()
            n = len(clean)
            max_s = int(fs * 30.0)
            if n > max_s: clean = clean[:max_s]; n = max_s
        else:
            dur = float(p.get("sig_dur_s", 2))
            clean = gen_emg(fs, dur, p["emg_amp"])
            n = len(clean)

        art, times, interval = gen_artefact(
            fs, n, p["stim_freq"], p["art_amp"],
            emg_amp=p["emg_amp"] * p["mwave_pct"])
        ch1 = clean + art + self._get_noise_segment(n, channel=1)

        if self.loaded_ch2 is not None:
            raw_ch2 = self.loaded_ch2[:n].copy()
            art_ch2 = art * (0.92 + np.random.randn() * 0.02)
            ch2 = raw_ch2 + art_ch2 + self._get_noise_segment(n, channel=2)
        else:
            ch2 = make_ch2(clean, art, fs, emg_amp=p["emg_amp"] * p["mwave_pct"])

        algo_keys = list(ALGO_NAMES.keys())
        mode_keys = [mid for _, mid, _ in METRIC_MODES]

        dispatch = {
            "blank"   : lambda: algo_blanking(ch1, times, fs, p["blank_ms"]),
            "template": lambda: algo_fixed_template(ch1, times, interval, fs, p["n_avg"]),
            "ewma"    : lambda: algo_ewma_template(ch1, times, interval, fs, p["ewma_alpha"]),
            "destd"   : lambda: algo_destd(ch1, ch2, times, interval, fs),
            "gso"     : lambda: algo_gso(ch1, ch2, times, interval, fs),
            "lms"     : lambda: algo_lms(ch1, ch2, times, interval, fs, p["lms_mu"]),
            "enlms"   : lambda: algo_enlms(ch1, ch2, times, interval, fs, p["lms_mu"]),
            "rls"     : lambda: algo_rls(ch1, ch2, times, interval, fs, p["rls_lam"]),
            "ceemdan" : lambda: algo_ceemdan(ch1, times, interval, fs, p["stim_freq"]),
        }

        algo_outputs = {}; algo_lats = {}
        for k in algo_keys:
            out, lat = dispatch[k]()
            algo_outputs[k] = out; algo_lats[k] = lat

        all_results = {}
        for k in algo_keys:
            all_results[k] = {}
            for mode in mode_keys:
                try:
                    m = compute_metrics(
                        clean, algo_outputs[k], k,
                        p["blank_ms"], p["stim_freq"],
                        times=times, fs=fs, mode=mode, art=art)
                    m["latency"] = algo_lats[k]
                except Exception:
                    m = {"latency": algo_lats[k]}
                all_results[k][mode] = m

        # ── style helpers ─────────────────────────────────────────────────────
        wb  = Workbook()
        ts  = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        dsrc = self.file_label.get()

        def _fill(h): return PatternFill("solid", fgColor=h)
        def _font(sz=9, bold=False, color="222222"):
            return Font(name="Calibri", size=sz, bold=bold, color=color)
        def _hdr(sz=9, color="FFFFFF"):
            return Font(name="Calibri", size=sz, bold=True, color=color)
        def _ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
        def _lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=False)
        def _bdr():
            s = Side(style="thin", color="CCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)

        AHX = {
            "blank":"6B7280","template":"94A3B8","ewma":"38BDF8",
            "destd":"6C8FFF","gso":"34D399","lms":"A78BFA",
            "enlms":"C084FC","rls":"F97316","ceemdan":"FB7185",
        }

        # ── Sheet 1: All Metrics ──────────────────────────────────────────────
        ws1 = wb.active; ws1.title = "All Metrics"
        ws1.sheet_view.showGridLines = False; ws1.freeze_panes = "D5"

        ws1.merge_cells("A1:L1")
        c = ws1["A1"]; c.value = f"EMG Artefact Suppression Benchmark  —  {ts}"
        c.font = _hdr(12); c.fill = _fill("0F1117"); c.alignment = _ctr()
        ws1.row_dimensions[1].height = 26

        ws1.merge_cells("A2:L2")
        c = ws1["A2"]
        c.value = (f"Data: {dsrc}  |  fs={fs}Hz  stim={p['stim_freq']}Hz  "
                   f"art={int(p['art_amp']*1000)}mV  EMG={int(p['emg_amp']*1e6)}µV  "
                   f"blank={p['blank_ms']}ms  n_avg={p['n_avg']}  "
                   f"alpha={p['ewma_alpha']:.2f}  mu={p['lms_mu']:.3f}  "
                   f"lam={p['rls_lam']:.3f}")
        c.font = _font(8, color="AAAAAA"); c.fill = _fill("181B27"); c.alignment = _lft()
        ws1.row_dimensions[2].height = 16; ws1.row_dimensions[3].height = 6

        hrow = ["Group","Metric","Unit","Target"] + [ALGO_NAMES[k] for k in algo_keys]
        cwid = [20,22,8,22] + [14]*len(algo_keys)
        hfil = ["1A1D2A"]*4 + [AHX[k] for k in algo_keys]
        for ci,(h,w,hf) in enumerate(zip(hrow,cwid,hfil),1):
            ws1.column_dimensions[get_column_letter(ci)].width = w
            c = ws1.cell(row=4, column=ci, value=h)
            c.font=_hdr(); c.fill=_fill(hf); c.alignment=_ctr(); c.border=_bdr()
        ws1.row_dimensions[4].height = 30

        MROWS = [
            ("Your §2.10",      "SNR (dB)",          "inter",   "snr",      "dB", "higher=better",        True),
            ("Your §2.10",      "SDR (dB)",          "inter",   "sdr",      "dB", "higher=better",        True),
            ("Your §2.10",      "RMSE (µV)",         "inter",   "rmse",     "µV", "lower=better",         False),
            ("Your §2.10",      "Pearson r",         "inter",   "r",        "",   "→ 1.0",                True),
            ("Your §2.10",      "Data loss (%)",     "inter",   "loss",     "%",  "0% (blank only)",      False),
            ("Your §2.10",      "Latency (ms)",      "inter",   "latency",  "ms", "lower=better",         False),
            ("Wang/Chen 21/23", "SNR full (dB)",     "wang",    "snr",      "dB", "−15 to −46 dB",        True),
            ("Wang/Chen 21/23", "NRMSE",             "wang",    "nrmse",    "",   "2.28–11.35",           False),
            ("Wang/Chen 21/23", "RMSE (µV)",         "wang",    "rmse",     "µV", "lower=better",         False),
            ("Wang/Chen 21/23", "Pearson r",         "wang",    "r",        "",   "→ 0.94",               True),
            ("Sennels 1997",    "MRI_y (dB)",        "sennels", "mri_y",    "dB", "→ 0 dB",              None),
            ("Sennels 1997",    "PR (dB)",           "sennels", "pr",       "dB", "14–28 dB",             True),
            ("Sennels 1997",    "RMSE (µV)",         "sennels", "rmse",     "µV", "lower=better",         False),
            ("Liu 2025 ASR",    "ASR (dB)",          "asr",     "asr",      "dB", "> 20 dB",              True),
            ("Liu 2025 ASR",    "RMSE (µV)",         "asr",     "rmse",     "µV", "lower=better",         False),
            ("Mandrile 2003",   "ARV_norm (%)",      "mandrile","arv_norm", "%",  "< 46%",                False),
            ("Mandrile 2003",   "RMSE (µV)",         "mandrile","rmse",     "µV", "lower=better",         False),
            ("Limnuson 2014",   "RMS reduction ×",   "limnuson","rms_red",  "×",  "> 5× / 17×",           True),
            ("Limnuson 2014",   "M-wave CC",         "limnuson","sdr",      "",   "0.64→0.89",            True),
            ("Limnuson 2014",   "M-wave RMSE (µV)",  "limnuson","rmse",     "µV", "425→223 µV",          False),
            ("Andrews 2023",    "FFT% vs input",     "andrews", "fft_pct",  "%",  "~96% (AA-IF)",         True),
            ("Andrews 2023",    "FFT% vs clean",     "andrews", "sdr",      "%",  "higher=better",        True),
        ]
        GBG = {
            "Your §2.10":"1E3A5F","Wang/Chen 21/23":"1A3A4A",
            "Sennels 1997":"2D2040","Liu 2025 ASR":"1A3A2A",
            "Mandrile 2003":"2D2D2D","Limnuson 2014":"3A2010","Andrews 2023":"3A1020",
        }
        prev_g = None
        for ri,(grp,metric,mkey,fld,unit,tgt,higher) in enumerate(MROWS, 5):
            ws1.row_dimensions[ri].height = 18
            bg = GBG.get(grp,"222222")
            c=ws1.cell(row=ri,column=1,value=grp if grp!=prev_g else "")
            c.font=_font(8,bold=True,color="CCCCCC"); c.fill=_fill(bg)
            c.alignment=_lft(); c.border=_bdr(); prev_g=grp
            c=ws1.cell(row=ri,column=2,value=metric)
            c.font=_font(9,color="DDDDDD"); c.fill=_fill("1C1F2C")
            c.alignment=_lft(); c.border=_bdr()
            c=ws1.cell(row=ri,column=3,value=unit)
            c.font=_font(8,color="888888"); c.fill=_fill("1C1F2C")
            c.alignment=_ctr(); c.border=_bdr()
            c=ws1.cell(row=ri,column=4,value=tgt)
            c.font=_font(8,color="6B9FBF"); c.fill=_fill("1A2030")
            c.alignment=_lft(); c.border=_bdr()
            vals=[all_results[k].get(mkey,{}).get(fld,None) for k in algo_keys]
            nums=[v for v in vals if isinstance(v,(int,float))]
            best=(max(nums) if higher else min(nums)) if nums and higher is not None else None
            for ci2,(k,v) in enumerate(zip(algo_keys,vals),5):
                c=ws1.cell(row=ri,column=ci2)
                if isinstance(v,(int,float)):
                    c.value=round(float(v),4); c.number_format="0.00"
                    ib=best is not None and abs(v-best)<1e-9
                    c.font=_font(10,bold=ib,color="FFFFFF" if ib else "DDDDDD")
                    c.fill=_fill(AHX[k] if ib else "1C1F2C")
                else:
                    c.value="—"; c.font=_font(9,color="555555"); c.fill=_fill("1C1F2C")
                c.alignment=_ctr(); c.border=_bdr()

        # ── Sheet 2: Raw Data ─────────────────────────────────────────────────
        ws2=wb.create_sheet("Raw Data"); ws2.sheet_view.showGridLines=False
        afl=["snr","sdr","rmse","r","loss","latency",
             "nrmse","mri_y","pr","asr","arv_norm","rms_red","fft_pct"]
        for ci,h in enumerate(["Algorithm","Mode"]+afl,1):
            ws2.column_dimensions[get_column_letter(ci)].width=14
            c=ws2.cell(row=1,column=ci,value=h)
            c.font=_hdr(); c.fill=_fill("1A1D2A"); c.alignment=_ctr(); c.border=_bdr()
        ws2.column_dimensions["A"].width=18; ws2.column_dimensions["B"].width=18
        rr=2
        for k in algo_keys:
            for mode in mode_keys:
                m=all_results[k].get(mode,{})
                rv=[ALGO_NAMES[k],mode]+[
                    round(float(m.get(f,0)),6) if isinstance(m.get(f),(int,float)) else "—"
                    for f in afl]
                for ci2,v in enumerate(rv,1):
                    c=ws2.cell(row=rr,column=ci2,value=v)
                    c.font=_font(9,color="DDDDDD")
                    c.fill=_fill("181B27" if rr%2==0 else "1C1F2C")
                    c.alignment=_ctr(); c.border=_bdr()
                ws2.row_dimensions[rr].height=16; rr+=1

        # ── Sheet 3: Literature Targets ───────────────────────────────────────
        ws3=wb.create_sheet("Literature Targets"); ws3.sheet_view.showGridLines=False
        for ci,(h,w) in enumerate(zip(
                ["Paper","Method","Metric","Value","Condition","Notes"],
                [18,20,18,20,24,34]),1):
            ws3.column_dimensions[get_column_letter(ci)].width=w
            c=ws3.cell(row=1,column=ci,value=h)
            c.font=_hdr(); c.fill=_fill("1A1D2A"); c.alignment=_ctr(); c.border=_bdr()
        LIT=[
            ("Limnuson 2014","EWMA Template","RMS reduction","17×","Aplysia","IIR alpha=1/16 FPGA"),
            ("Limnuson 2014","EWMA Template","RMS reduction","5.3×","Rat cortex","IIR alpha=1/16 FPGA"),
            ("Liu 2025","Fixed Template","ASR","20.2 dB","Saline phantom","Randles model ASIC"),
            ("Liu 2025","Pole-shifting","ASR","~6 dB","Baseline","Simple hardware"),
            ("Wang 2021","GS-APEF/LMS-AF","SNR","−15.26±3.87 dB","r=4","Full signal"),
            ("Wang 2021","GS-APEF/LMS-AF","SNR","−46.19±6.53 dB","r=20","Full signal"),
            ("Wang 2021","GS-APEF/LMS-AF","NRMSE","2.28±0.36","r=4",""),
            ("Wang 2021","GS-APEF/LMS-AF","NRMSE","11.35±2.22","r=20",""),
            ("Wang 2021","GS-APEF vs LMS","p-value","p > 0.05","Two-way ANOVA","No significant diff"),
            ("Chen 2023","DESTD","NRMSE","0.0576±0.0187","20 mA","Best performance"),
            ("Chen 2023","DESTD","Robustness","p > 0.05","All currents","Stable as I increases"),
            ("Sennels 1997","Adaptive M=1","MRI_y","0.1 dB","No variation","Near 0 dB target"),
            ("Sennels 1997","Adaptive M=6","MRI_y","−7.5 to −2.4 dB","±100% variation","Worst case"),
            ("Sennels 1997","Adaptive M=1-6","PR","14–28 dB","Real recordings","Power reduction"),
            ("Mandrile 2003","Various","ARV_norm","46–91%","Electrode distance","Lower=better"),
            ("Liu 2014","Savitzky-Golay","M-wave CC","0.89","After suppression","vs 0.64 before"),
            ("Liu 2014","Savitzky-Golay","M-wave RMSE","223 µV","After suppression","vs 425 µV before"),
            ("Andrews 2023","AA-IF","FFT% preserved","96±5%","vs contaminated","AA-IF method"),
            ("Andrews 2023","EMD-BF","FFT% preserved","75±6%","vs contaminated","Baseline"),
            ("Andrews 2023","AA-IF vs EMD-BF","Significance","p < 0.001","—","AA-IF better"),
        ]
        PBG={"Limnuson 2014":"3A2010","Liu 2025":"1A3A2A","Wang 2021":"1A3A4A",
             "Chen 2023":"1A2A3A","Sennels 1997":"2D2040","Mandrile 2003":"2D2D2D",
             "Liu 2014":"1A3A2A","Andrews 2023":"3A1020"}
        for ri3,rd in enumerate(LIT,2):
            bg=PBG.get(rd[0],"222222"); ws3.row_dimensions[ri3].height=16
            for ci3,v in enumerate(rd,1):
                c=ws3.cell(row=ri3,column=ci3,value=v)
                c.font=_font(9,color="DDDDDD"); c.fill=_fill(bg)
                c.alignment=_lft(); c.border=_bdr()

        # ── Sheet 4: Params ───────────────────────────────────────────────────
        ws4=wb.create_sheet("Params"); ws4.sheet_view.showGridLines=False
        ws4.column_dimensions["A"].width=28; ws4.column_dimensions["B"].width=22
        prows=[
            ("Export timestamp",ts),("Data source",dsrc),
            ("fs (Hz)",fs),("Stim freq (Hz)",p["stim_freq"]),
            ("Artefact amp (mV)",round(p["art_amp"]*1000,1)),("EMG amp (µV RMS)",round(p["emg_amp"]*1e6,1)),
            ("M-wave scale (%)",round(p["mwave_pct"]*100,1)),("Blank window (ms)",p["blank_ms"]),
            ("Template avg cycles",p["n_avg"]),("EWMA alpha",round(p["ewma_alpha"],4)),
            ("LMS/eNLMS mu",round(p["lms_mu"],4)),("RLS lambda",round(p["rls_lam"],4)),
            ("Signal length (s)",n/fs),
        ]
        for ri4,(k,v) in enumerate(prows,1):
            ws4.row_dimensions[ri4].height=18
            c=ws4.cell(row=ri4,column=1,value=k)
            c.font=_font(9,color="AAAAAA"); c.fill=_fill("1C1F2C")
            c.alignment=_lft(); c.border=_bdr()
            c=ws4.cell(row=ri4,column=2,value=v)
            c.font=_font(10,bold=True,color="6C8FFF"); c.fill=_fill("181B27")
            c.alignment=_ctr(); c.border=_bdr()

        wb.save(path)


# =============================================================================

if __name__ == "__main__":
    np.random.seed(42)
    App().mainloop()